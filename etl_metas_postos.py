#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ETL Metas Postos — parseia METAS POSTO {ano}.xlsx (abas numéricas 1..12 = mês)
e gera dados_dre_postos_adaptive/metas_postos_{ano}.json com a META de cada
indicador por posto/mês.

Fonte: \\10.61.1.13\\controller\\03 - POSTOS\\METAS POSTO {ano}.xlsx — abas "1".."12"
(uma por mês), com header na linha 4 (índice 3). Posto na col 1 (P01..P11), e
colunas Meta: Volume(2), Margem %(8), Perdas e Sobras(12), Nº Abast(16),
Ticket Médio R$(21).

As metas são por POSTO e por MÊS. O REALIZADO vem das queries SQL (não daqui).

Uso: python3 etl_metas_postos.py [caminho.xlsx] [ano]
"""
import os, sys, json, glob
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(BASE, "dados_dre_postos_adaptive")
META_DIRS = ["/mnt/controller/03 - POSTOS", "/mnt/c/Users/wesley/Downloads"]

# O layout das colunas MUDA entre abas (algumas têm coluna "Ating %" extra, que
# desloca tudo à direita). Por isso detectamos as colunas pelo NOME do cabeçalho
# (linha 4 / índice 3), não por índice fixo — senão pegávamos "Real" no lugar de
# "Meta" nos meses sem a coluna "Ating %". Mapa: campo → predicado sobre o header.
HEADER_ROW = 3
def _norm(s):
    return " ".join(str(s or "").replace("\n", " ").split()).strip().lower()

CAMPO_MATCH = {
    "volume": lambda h: h.startswith("meta volume"),
    "margem": lambda h: h.startswith("meta margem"),
    "perdas": lambda h: h.startswith("meta perdas"),
    "abast":  lambda h: h.startswith("meta") and "abast" in h,
    "ticket": lambda h: h.startswith("meta") and "ticket" in h,
}
# P01..P11 → código do posto no DADOS (001..011)
def _posto_cod(p):
    p = str(p).strip().upper()
    if p.startswith("P") and p[1:].isdigit():
        return p[1:].zfill(3)
    return None


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def achar_arquivo(ano):
    for d in META_DIRS:
        for c in glob.glob(os.path.join(d, f"METAS POSTO {ano}.xlsx")):
            if not os.path.basename(c).startswith("~$"):
                return c
    return None


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    metas = {}   # cod -> { "1": {volume,...}, ... }
    for aba in wb.sheetnames:
        if not aba.strip().isdigit():
            continue
        mes = int(aba.strip())
        if mes < 1 or mes > 12:
            continue
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= HEADER_ROW:
            continue
        hdr = [_norm(c) for c in rows[HEADER_ROW]]
        # Resolve a coluna de cada campo pelo nome do cabeçalho desta aba.
        colmap = {}
        for ci, h in enumerate(hdr):
            for campo, pred in CAMPO_MATCH.items():
                if campo not in colmap and pred(h):
                    colmap[campo] = ci
        # Coluna do posto (header "posto"/"postos"; senão assume col 1).
        col_posto = next((ci for ci, h in enumerate(hdr) if h in ("posto", "postos")), 1)
        if not colmap:
            continue
        for r in rows[HEADER_ROW + 1:]:
            if not r or len(r) <= col_posto:
                continue
            cod = _posto_cod(r[col_posto])
            if not cod:
                continue
            ind = {}
            for campo, ci in colmap.items():
                if ci < len(r):
                    v = _num(r[ci])
                    if v is not None:
                        ind[campo] = v
            if ind:
                metas.setdefault(cod, {})[str(mes)] = ind
    wb.close()
    return metas


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else None
    ano = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    if not path:
        path = achar_arquivo(ano)
    if not path or not os.path.isfile(path):
        print(f"  · (sem METAS POSTO {ano}.xlsx — pulando)")
        return
    print("  fonte:", path)
    metas = parse(path)
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, f"metas_postos_{ano}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump({"ano": ano, "metas": metas}, f, ensure_ascii=False, separators=(",", ":"))
    npost = len(metas)
    nmes = sum(len(m) for m in metas.values())
    print(f"  ✓ {out} ({npost} postos · {nmes} posto-mês)")


if __name__ == "__main__":
    main()
