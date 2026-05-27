"""ETL de Energia · lê direto as abas mensais de
   '2026/Contas de Energia - SUPERMERCADOS.xlsx' e gera energia.json.

Abas processadas (cada uma vira um "mês" no resumo):
  JAN, FEV, MAR, ABR, MAI, JUN, JUL, AGO, SET, OUT, NOV, DEZ  (2025)
  JAN.26, FEV.26, MAR.26, ABR.26, MAI.26                       (2026)

Estrutura de cada aba mensal:
  Row 1: título "CONTROLE DE ENERGIA - SUPERMERCADOS"
  Row 2: cabeçalho — Loja, Mês, Nº Inscrição, ..., Total, % S/Venda,
         Venc Fornec, Qtd dias, Leitura, Venc GD/ML, TOTAL KWh, INJEÇÃO,
         kWh FORNEC R$, kWh GD/ML R$, FORNEC + GD/ML (R$/kWh),
         (%) SOBRE INJEÇÃO[, ENCARGO, DATA  · ABR em diante]
  Row 3+: linhas Lxx (válidas) + linhas-subtotal (Loja vazia) +
          linhas "OBSERVAÇÃO" → filtramos só as Lxx.

Vendas vêm da aba Vda_Super (Loja, Mês-serial, Valor, MêsTexto).

Saída: energia.json com lojas[], meses[], registros[], status_ml_gd[],
contas_abertas[], situacao[], class_status[].
"""
from __future__ import annotations
import datetime as _dt
import json
import math
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl não instalado · pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path("/mnt/controller/02 - SUPERMERCADOS/98 - CONTAS DE ENERGIA")
CONTAS = ROOT / "2026" / "Contas de Energia - SUPERMERCADOS.xlsx"
OUT = Path("/root/projeto_dre/energia.json")

# Ordem cronológica das abas mensais
MESES = [
    "JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ",
    "JAN.26", "FEV.26", "MAR.26", "ABR.26", "MAI.26",
]

LOJA_RE = re.compile(r"^L\d+[a-z]?$", re.IGNORECASE)


def classificar_mes(label: str):
    """'MAI.26' → {'desc_antiga': 'MAI.26', 'ano': 2026, 'mes': 'MAI'}.
    'JAN'     → {'desc_antiga': 'JAN',     'ano': 2025, 'mes': 'JAN'}.
    Convenção: sufixo .YY indica ano 20YY; sem sufixo é 2025 (base atual)."""
    m = re.match(r"^([A-Z]{3})(?:\.(\d{2}))?$", label, re.IGNORECASE)
    if m:
        mes = m.group(1).upper()
        ano = 2000 + int(m.group(2)) if m.group(2) else 2025
        return {"desc_antiga": label, "ano": ano, "mes": mes}
    return {"desc_antiga": label, "ano": 2025, "mes": label}


def _f(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    except (TypeError, ValueError):
        return None


def _s(v):
    if v is None:
        return ""
    return str(v).strip()


def _excel_date(v):
    """Excel serial → YYYY-MM-DD. Aceita também datetime, time, string."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            base = _dt.datetime(1899, 12, 30)
            return (base + _dt.timedelta(days=float(v))).strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(v, _dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, _dt.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, _dt.time):
        return None  # time(0,0) sentinela do Excel
    return str(v)


def _normaliza_loja(loja: str) -> str:
    """L01a, L01b, L01A → L01 (agrupa UCs da mesma loja)."""
    m = re.match(r"^(L\d+)", loja, re.IGNORECASE)
    return m.group(1).upper() if m else loja.upper()


def read_mes_tab(wb, sheet_name):
    """Lê uma aba mensal e devolve lista de registros (1 por UC)."""
    if sheet_name not in wb.sheetnames:
        print(f"  ! aba {sheet_name} não encontrada — pulando")
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True, max_col=40)
    # row 1 é título; row 2 é header
    try:
        _row1 = next(rows_iter)
        header = next(rows_iter)
    except StopIteration:
        return []

    idx = {(_s(h) if h else ""): i for i, h in enumerate(header)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    cls = classificar_mes(sheet_name)
    for row in rows_iter:
        loja_raw = _s(col(row, "Loja"))
        if not LOJA_RE.match(loja_raw):
            continue  # subtotal ou OBSERVAÇÃO
        out.append({
            "loja": _normaliza_loja(loja_raw),
            "loja_uc": loja_raw,
            "mes": sheet_name,           # mantém DESCANTIGA pra compat
            "desc_antiga": cls["desc_antiga"],
            "ano": cls["ano"],
            "mes_curto": cls["mes"],
            "uc": _s(col(row, "Nº Inscrição")),
            "ponta": _f(col(row, "Ponta")),
            "fora_ponta": _f(col(row, "Fora Ponta")),
            "cativo": _f(col(row, "Valor Cativo R$")),
            "dem": _f(col(row, "Dem")),
            "ult_dem": _f(col(row, "Ult Dem")),
            "multas": _f(col(row, "Multas e Juros")),
            "taxas": _f(col(row, "Taxas Diversas")),
            "valor_st": _f(col(row, "Valor s/ Taxas")),
            "valor_inj": _f(col(row, "Valor c/ Injeção")),
            "gd_ml": _f(col(row, "Valor GD/ML")),
            "desconto": _f(col(row, "Desconto")),
            "desc_pct": _f(col(row, "Desc (%)")),
            "total": _f(col(row, "Total")),
            "pct_venda": _f(col(row, "(%) S/ Venda")),
            "venc_fornec": _excel_date(col(row, "Venc Fornec")),
            "qtd_dias": _f(col(row, "Qtd dias")),
            "leitura": _excel_date(col(row, "Leitura")),
            "venc_gd": _excel_date(col(row, "Venc GD/ML")),
            "kwh": _f(col(row, "TOTAL KWh")),
            "injecao": _f(col(row, "INJEÇÃO")),
            "kwh_fornec_rs": _f(col(row, "kWh FORNEC R$")),
            "kwh_gd_rs": _f(col(row, "kWh GD/ML R$")),
            "rs_kwh": _f(col(row, "FORNEC + GD/ML (R$/kWh)")),
            "pct_inj": _f(col(row, "(%) SOBRE INJEÇÃO")),
            "encargo": _f(col(row, "ENCARGO")),
        })
    return out


def read_vda_super(wb):
    """Vda_Super → {(loja, mes_texto): valor}. mes_texto = 'JAN', 'FEV', etc.
    NOTA: a aba não tem o sufixo '.26' — distinguimos pelo Mês-serial:
    serial < 45658 + 365 = 46023 → 2025, senão → 2026."""
    if "Vda_Super" not in wb.sheetnames:
        return {}
    ws = wb["Vda_Super"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    map_ = {}
    for row in rows_iter:
        loja = _s(row[0]) if row[0] else ""
        if not LOJA_RE.match(loja):
            continue
        valor = _f(row[2])
        mes_txt = _s(row[3]) if len(row) > 3 else ""
        # row[1] pode vir como datetime.datetime ou serial numérico
        m1 = row[1]
        ano = None
        if isinstance(m1, _dt.datetime):
            ano = m1.year
        elif isinstance(m1, _dt.date):
            ano = m1.year
        elif isinstance(m1, (int, float)) and m1:
            # serial → ano
            try:
                ano = (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(m1))).year
            except Exception:
                ano = None
        if not ano or not mes_txt:
            continue
        sufixo = ".26" if ano >= 2026 else ""
        chave = (_normaliza_loja(loja), mes_txt + sufixo)
        map_[chave] = (map_.get(chave) or 0) + (valor or 0)
    return map_


def read_aux(wb, sheet_name):
    """Lê uma aba auxiliar (CONTAS_ABERTO, ML-GD) como lista de dicts."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [(_s(h) if h else "") for h in header]
    out = []
    for row in rows_iter:
        loja = _s(row[0]) if row else ""
        if not loja:
            continue
        d = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            # Datas — colunas que começam com DTA
            if h.startswith("DTA"):
                d[h] = _excel_date(v)
            elif h in ("VLR VENCIDO", "VLR ATUAL", "VLR FUTURA"):
                d[h] = _f(v)
            else:
                d[h] = _s(v)
        out.append(d)
    return out


def read_situacao(wb):
    """Aba 'Situação' — descrição executiva por loja."""
    if "Situação" not in wb.sheetnames:
        return []
    ws = wb["Situação"]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [(_s(h) if h else "") for h in header]
    out = []
    for row in rows_iter:
        if not row or not _s(row[0]):
            continue
        loja = _s(row[0])
        if not LOJA_RE.match(loja):
            continue
        d = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            d[h] = _s(v) if isinstance(v, str) or v is None else v
        out.append(d)
    return out


def read_class_status(wb):
    """Aba 'Class' — status mensal de cada UC (FEV, MAR, ABR, MAI)."""
    if "Class" not in wb.sheetnames:
        return []
    ws = wb["Class"]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [(_s(h) if h else "") for h in header]
    out = []
    for row in rows_iter:
        if not row or not _s(row[0]):
            continue
        d = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            if isinstance(v, (int, float)) and v > 30000 and v < 80000:
                # Excel serial date
                d[h] = _excel_date(v)
            else:
                d[h] = _s(v) if v is not None else ""
        out.append(d)
    return out


def main():
    print(f"lendo {CONTAS}")
    wb = load_workbook(CONTAS, data_only=True, read_only=True)

    # Vendas
    vendas = read_vda_super(wb)
    print(f"  · Vda_Super → {len(vendas)} pares loja-mês")

    # Registros das abas mensais. Venda é atribuída só ao primeiro UC de
    # cada (loja, mês) pra não duplicar quando o front somar por loja-mês.
    todos_registros = []
    lojas = set()
    meses_presentes = []
    for m in MESES:
        regs = read_mes_tab(wb, m)
        vendas_atribuidas = set()
        for r in regs:
            chave = (r["loja"], r["mes"])
            if chave not in vendas_atribuidas:
                r["venda"] = vendas.get(chave)
                vendas_atribuidas.add(chave)
            else:
                r["venda"] = None
            todos_registros.append(r)
            lojas.add(r["loja"])
        if regs:
            meses_presentes.append(m)
        print(f"  · {m:8s} → {len(regs)} registros")

    # Aux: ML-GD, CONTAS_ABERTO, Situação, Class
    ml_gd_raw = read_aux(wb, "ML-GD")
    abertos_raw = read_aux(wb, "CONTAS_ABERTO")

    def _norm(d, fields):
        """Normaliza chaves do aux pra forma compacta consumida pelo front."""
        return {
            "loja": d.get("LOJAS", ""),
            "uc": d.get("UC", ""),
            "cnpj": d.get("CNPJ", ""),
            **fields,
            "vlr_vencido": d.get("VLR VENCIDO"),
            "dt_venc": d.get("DTA VENC"),
            "vlr_atual": d.get("VLR ATUAL"),
            "dt_atual": d.get("DTA ATUAL"),
            "vlr_futura": d.get("VLR FUTURA"),
            "dt_futura": d.get("DTA FUTURA"),
            "obs": d.get("OBSERVAÇÃO", ""),
        }

    status_ml_gd = [_norm(d, {"empresa": d.get("EMPRESA", "")}) for d in ml_gd_raw]
    contas_abertas = [_norm(d, {}) for d in abertos_raw]
    situacao = read_situacao(wb)
    class_status = read_class_status(wb)

    # Classificação dos meses (DESCANTIGA|ANO|MÊS)
    classificacao = [classificar_mes(m) for m in meses_presentes]
    anos_distintos = sorted({c["ano"] for c in classificacao})

    out = {
        "geradoEm": _dt.datetime.now().isoformat(timespec="seconds"),
        "fonte": str(CONTAS),
        "lojas": sorted(lojas),
        "meses": meses_presentes,
        "classificacao": classificacao,   # [{desc_antiga, ano, mes}]
        "anos": anos_distintos,
        "registros": todos_registros,
        "status_ml_gd": status_ml_gd,
        "contas_abertas": contas_abertas,
        "situacao": situacao,
        "class_status": class_status,
    }

    OUT.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")))
    sz = OUT.stat().st_size / 1024
    print()
    print(f"  · {len(todos_registros)} registros · {len(lojas)} lojas · {len(meses_presentes)} meses")
    print(f"  · {len(status_ml_gd)} entradas ML-GD · {len(contas_abertas)} contas em aberto")
    print(f"  · {len(situacao)} lojas em Situação · {len(class_status)} UCs em Class")
    print(f"  · {OUT} ({sz:,.1f} KB)")


if __name__ == "__main__":
    main()
