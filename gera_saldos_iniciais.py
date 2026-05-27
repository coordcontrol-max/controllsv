"""Extrai SALDO INICIAL dos segmentos Postos e Outras a partir das planilhas
e gera 2 arquivos JSON pequenos consumidos pelo dashboard:

  /root/projeto_dre/dados_fluxo_postos/saldos_iniciais.json
  /root/projeto_dre/dados_fluxo_outras/saldos_iniciais.json

Estrutura:
  {"YYYY-MM": {"total": <float>, "porLoja": {"P01": <float>, ...}}, ...}

Fontes:
  - F2 abas DFC01..DFC12: linha "( = ) SALDO INICIAL" por loja (cols D-N postos,
    cols Q-U outras). Cobre meses históricos que têm aba DFCxx preenchida.
  - F1 RESUMO_Mês_Atual + F3 RESUMO: bloco "Saldo Inicial (Bancário)" do primeiro
    dia útil do mês (pega o dia 1; se ausente, pega o menor dia disponível).
    Cobre o mês atual (e às vezes Próximo_Mês).

Postos lojas = P01..P11 (col label exato). Outras = FLUXO, LP, PEGUI, RETA, TARES.
"""
from __future__ import annotations
import os, json, datetime as dt
from openpyxl import load_workbook

BASE = "/mnt/c/Users/wesley/Desktop/Postos e outras empresas"
F1 = os.path.join(BASE, "01 - Fluxo de Caixa Diário - Postos.xlsx")
F2 = os.path.join(BASE, "02 - Fluxo de Caixa - Postos e Outras Empresas.xlsx")
F3 = os.path.join(BASE, "02 - Fluxo de Caixa Diário - Outras Empresas (version 1).xlsx")

OUT_POSTOS = "/root/projeto_dre/dados_fluxo_postos"
OUT_OUTRAS = "/root/projeto_dre/dados_fluxo_outras"

POSTOS = [f"P{i:02d}" for i in range(1, 12)]
OUTRAS = ["FLUXO", "LP", "PEGUI", "RETA", "TARES"]


def parse_dfc0x(path, sheet):
    """Lê uma aba DFCxx. Retorna (ano, mes, postos_dict, outras_dict) onde os
    dicts mapeiam loja → saldo inicial. Linha "( = ) SALDO INICIAL" é row index 3."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return None, None, {}, {}
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            return None, None, {}, {}
        # row 1 col C tem a data de referência
        date_cell = rows[1][2] if len(rows[1]) > 2 else None
        ano = mes = None
        if isinstance(date_cell, dt.datetime):
            ano, mes = date_cell.year, date_cell.month
        header = rows[2]
        col_loja = {}
        for ci, c in enumerate(header):
            if not c: continue
            s = str(c).strip()
            if s in POSTOS or s in OUTRAS:
                col_loja[ci] = s
        # Procura a linha "( = ) SALDO INICIAL" — deve ser row 3, mas faz busca
        # genérica pra robustez.
        saldo_row = None
        for ri in range(2, min(15, len(rows))):
            lab = rows[ri][2] if len(rows[ri]) > 2 else None
            if lab and isinstance(lab, str) and "SALDO INICIAL" in lab.upper():
                saldo_row = rows[ri]
                break
        if not saldo_row:
            return ano, mes, {}, {}
        postos_out = {}
        outras_out = {}
        for ci, loja in col_loja.items():
            if ci >= len(saldo_row): continue
            v = saldo_row[ci]
            if v is None: continue
            try: v = float(v)
            except (TypeError, ValueError): continue
            if loja in POSTOS: postos_out[loja] = round(v, 2)
            else:              outras_out[loja] = round(v, 2)
        return ano, mes, postos_out, outras_out
    finally:
        wb.close()


def parse_resumo_saldo_inicial(path, sheet):
    """Lê aba RESUMO. Procura a linha "Saldo Inicial (Bancário)" e pega o valor
    do primeiro dia útil do mês (menor dia presente). Retorna (ano, mes, saldo_total).
    Não tem breakdown por loja nesses arquivos."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return None, None, None
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # Encontra linha das datas
        date_row_idx = None
        for i, r in enumerate(rows[:12]):
            dts = [c for c in r if isinstance(c, dt.datetime)]
            if len(dts) >= 5:
                date_row_idx = i; break
        if date_row_idx is None:
            return None, None, None
        date_row = rows[date_row_idx]
        col_to_date = {ci: c.date() for ci, c in enumerate(date_row) if isinstance(c, dt.datetime)}
        if not col_to_date:
            return None, None, None
        ano, mes = next(iter(col_to_date.values())).year, next(iter(col_to_date.values())).month
        # Acha linha "Saldo Inicial (Bancário)"
        target_row = None
        for r in rows[date_row_idx+1:date_row_idx+150]:
            for c in r[:5]:
                if c and isinstance(c, str) and "SALDO INICIAL" in c.upper() and "BANC" in c.upper():
                    target_row = r; break
            if target_row: break
        if not target_row:
            return ano, mes, None
        # Pega o menor dia (geralmente col que bate com dia 1)
        primeiros = sorted(col_to_date.items(), key=lambda x: x[1])
        for ci, _ in primeiros:
            if ci < len(target_row) and target_row[ci] is not None:
                try:
                    return ano, mes, round(float(target_row[ci]), 2)
                except (TypeError, ValueError):
                    continue
        return ano, mes, None
    finally:
        wb.close()


def main():
    postos_saldos = {}
    outras_saldos = {}

    print("→ Lendo DFC01..DFC12 (meses históricos)…")
    for i in range(1, 13):
        sn = f"DFC{i:02d}"
        ano, mes, p, o = parse_dfc0x(F2, sn)
        if ano is None or mes is None:
            print(f"   {sn}: sem dados")
            continue
        chave = f"{ano:04d}-{mes:02d}"
        if p:
            postos_saldos[chave] = {
                "total": round(sum(p.values()), 2),
                "porLoja": p,
            }
            print(f"   {sn} ({chave}) postos: total={postos_saldos[chave]['total']:,.2f}  lojas={list(p)}")
        if o:
            outras_saldos[chave] = {
                "total": round(sum(o.values()), 2),
                "porLoja": o,
            }
            print(f"   {sn} ({chave}) outras: total={outras_saldos[chave]['total']:,.2f}  lojas={list(o)}")

    print("→ Lendo RESUMO_Mês_Atual de F1/F3 (mês atual)…")
    for path, sheet, dest, label in [
        (F1, "RESUMO_Mês_Atual",   postos_saldos, "F1 postos"),
        (F1, "RESUMO_Próximo_Mês", postos_saldos, "F1 postos próximo"),
        (F3, "RESUMO",             outras_saldos, "F3 outras"),
    ]:
        try:
            ano, mes, saldo = parse_resumo_saldo_inicial(path, sheet)
            if ano and mes and saldo is not None:
                chave = f"{ano:04d}-{mes:02d}"
                if chave in dest:
                    # Mantém o que já veio do DFC0X (tem breakdown por loja)
                    print(f"   {label} ({chave}): R$ {saldo:,.2f} — já tinha do DFC0X, mantendo")
                else:
                    dest[chave] = {"total": saldo, "porLoja": {}}
                    print(f"   {label} ({chave}): total=R$ {saldo:,.2f} (sem breakdown por loja)")
            else:
                print(f"   {label}: sem dados")
        except Exception as e:
            print(f"   {label}: ERRO {e}")

    os.makedirs(OUT_POSTOS, exist_ok=True)
    os.makedirs(OUT_OUTRAS, exist_ok=True)
    out_p = os.path.join(OUT_POSTOS, "saldos_iniciais.json")
    out_o = os.path.join(OUT_OUTRAS, "saldos_iniciais.json")
    with open(out_p, "w", encoding="utf-8") as f:
        json.dump(postos_saldos, f, ensure_ascii=False, indent=2)
    with open(out_o, "w", encoding="utf-8") as f:
        json.dump(outras_saldos, f, ensure_ascii=False, indent=2)
    print(f"\nOK")
    print(f"  {out_p}  ({len(postos_saldos)} meses)")
    print(f"  {out_o}  ({len(outras_saldos)} meses)")


if __name__ == "__main__":
    main()
