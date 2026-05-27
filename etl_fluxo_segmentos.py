"""ETL — Fluxo de Caixa de POSTOS e OUTRAS EMPRESAS  (v2)

Cada segmento tem seu PRÓPRIO plano de contas. O canônico é extraído
da aba ANALÍTICO do arquivo específico de cada segmento (que define as
seções ENTRADAS/SAÍDAS/FINANCIAMENTO/INVESTIMENTO e suas linhas).

  • Postos canônico  = File 1 "ANALÍTICO_Mês_Atual"   (88 linhas)
  • Outras canônico  = File 3 "ANALÍTICO"             (35 linhas)

Fontes de dados:
  • File 1 (Postos): diário consolidado Abr/Mai 2026
  • File 2 (DFC01-03): mensal por loja Jan/Fev/Mar 2026 (postos e outras juntos)
  • File 3 (Outras): diário consolidado Abr 2026

Filtragem:
  Qualquer label da planilha que não esteja no plano de contas canônico
  do segmento (após aliases) é DESCARTADO. Isto cobre:
    - agregadores: ( + ) RECEBIMENTOS, ( = ) SALDO FINAL, % Despesas, ...
    - sub-credores: AMEX, ELO CREDITO, MASTERCARD CREDITO, ...
    - linhas que só pertencem ao outro segmento (cross-contamination)
"""
from __future__ import annotations
import os
import json
import datetime as dt
import calendar
from collections import defaultdict
from openpyxl import load_workbook

BASE = "/mnt/c/Users/wesley/Desktop/Postos e outras empresas"
F1 = os.path.join(BASE, "01 - Fluxo de Caixa Diário - Postos.xlsx")
F2 = os.path.join(BASE, "02 - Fluxo de Caixa - Postos e Outras Empresas.xlsx")
F3 = os.path.join(BASE, "02 - Fluxo de Caixa Diário - Outras Empresas (version 1).xlsx")
F4_TRANSF_POSTOS = os.path.join(BASE, "transferências bancárias - postos.xlsx")

OUT_POSTOS = "/root/projeto_dre/dados_fluxo_postos"
OUT_OUTRAS = "/root/projeto_dre/dados_fluxo_outras"

POSTOS = [f"P{i:02d}" for i in range(1, 12)]
OUTRAS = ["FLUXO", "LP", "PEGUI", "RETA", "TARES"]

# Estrutura fixa de grupos (compartilhada com Supermercados a partir de 18/05/2026).
# Antes era ["ATIVIDADES OPERACIONAIS","ATIVIDADES OPERACIONAIS","ATIVIDADES DE FINANCIAMENTO","Investimento"] — agora consolida
# Entradas+Saídas em ATIVIDADES OPERACIONAIS pra bater com o layout do Supermer.
GRUPOS = ["ATIVIDADES OPERACIONAIS", "ATIVIDADES DE FINANCIAMENTO", "ATIVIDADES DE INVESTIMENTO"]

# Mapa seção-header → (grupo, agrupamento). Linhas específicas dentro de
# OPERACIONAIS são re-categorizadas via LINHA_AGRUPAMENTO_OVERRIDE (Fornecedores
# de Mercadorias → "Fornecedores"; Pagto de Compra Entre Unidades → "Pagto
# Entre Unidades"; restante = "Despesas").
SECTION_KEYWORDS = {
    "ENTRADAS":                    ("ATIVIDADES OPERACIONAIS",  "Recebimentos Operacionais"),
    "RECEBIMENTOS":                ("ATIVIDADES OPERACIONAIS",  "Recebimentos Operacionais"),
    "SAÍDAS":                      ("ATIVIDADES OPERACIONAIS",  "Despesas"),
    "FORNECEDORES":                ("ATIVIDADES OPERACIONAIS",  "Fornecedores"),
    "DESPESAS":                    ("ATIVIDADES OPERACIONAIS",  "Despesas"),
    "ATIVIDADES DE FINANCIAMENTO": ("ATIVIDADES DE FINANCIAMENTO", "Atividades de Financiamento"),
    "ATIVIDADES DE INVESTIMENTO":  ("ATIVIDADES DE INVESTIMENTO",  "Atividades de Investimento"),
}
LINHA_AGRUPAMENTO_OVERRIDE = {
    "Fornecedores de Mercadorias":    "Fornecedores",
    "Pagto de Compra Entre Unidades": "Pagto Entre Unidades",
}

# Aliases por segmento: variações encontradas no DFC01/02/03 (File 2) que
# precisam ser normalizadas pro nome canônico do segmento.
# Importante: o DFC01/02 condensa Fornecedores e Pagto Entre Unidades em
# linhas de uma palavra (que coincidem com nomes de agrupamentos). Sem alias,
# elas seriam descartadas pelo filtro e os valores de Fornecedores ficariam
# zerados nos meses Jan/Fev/Mar.
ALIAS = {
    "postos": {
        "Recbto de Venda em Crédito":            "Recebmto de Venda em Crédito",
        "Recbto de Venda em Débito":             "Recebmto de Venda em Débito",
        "Recbto de Venda em Crédito (Frotas)":   "Recebmto de Venda em Crédito (Frotas)",
        "Recbto de Venda em Dinheiro (Protege)": "Recebmto de Venda em Dinheiro (Protege)",
        "Recbto de Venda em PIX":                "Recebmto de Venda em PIX",
        "Recbto de Venda Entre Unidades":        "Recebmto de Venda Entre Unidades",
        "Recbto de Venda A Prazo":               "Recebmto de Vendas a Prazo",
        # DFC01/02 colapsa "Fornecedores de Mercadorias" em "Fornecedores"
        "Fornecedores":                          "Fornecedores de Mercadorias",
        "Pagmto Compra Entre Unidades":          "Pagto de Compra Entre Unidades",
        # Financiamento — File 2 usa nomes diferentes
        "Mútuos Recebidos":                      "Mútuos a Receber",
        "Mútuos Recebidos Entre Grupos":         "Mútuos a Receber (Entre grupos)",
    },
    "outras": {
        "Aluguéis Recebidos":                    "Recebmto de Aluguéis",
        "Fgts":                                  "FGTS",
        "Inss":                                  "INSS",
        "Mútuos Recebidos":                      "Mútuos a Receber",
        "Mútuos Recebidos Entre Grupos":         "Mútuos a Receber (Entre Grupos)",
        "Mútuo a Pagar (Entre Grupos)":          "Mútuo a Pagar (Entre Grupos)",
        # Variações plural/singular do File 2
        "Despesas Bancaria":                     "Despesa Bancária",
        # PEGUI/TARES também vendem combustível/produtos → têm PIX
        "Recbto de Venda em PIX":                "Recebmto de Venda em PIX",
        "Recbto de Venda em Crédito":            "Recebmto de Venda em Crédito",
        "Recbto de Venda em Débito":             "Recebmto de Venda em Débito",
        "Recbto de Venda em Dinheiro (Protege)": "Recebmto de Venda em Dinheiro (Protege)",
    },
}

# Linhas adicionais que aparecem nos DFC0X de OUTRAS (PEGUI/TARES principalmente)
# mas não estão no canonical extraído do File 3 ANALÍTICO. São despesas válidas
# que precisam ser preservadas no plano de contas de outras.
LINHAS_OUTRAS_EXTRAS = [
    # (linha, grupo, agrupamento)
    ("Recebmto de Venda em Crédito",       "ATIVIDADES OPERACIONAIS", "Recebimentos Operacionais"),
    ("Recebmto de Venda em Débito",        "ATIVIDADES OPERACIONAIS", "Recebimentos Operacionais"),
    ("Recebmto de Venda em Dinheiro (Protege)", "ATIVIDADES OPERACIONAIS", "Recebimentos Operacionais"),
    ("Recebmto de Venda em PIX",           "ATIVIDADES OPERACIONAIS", "Recebimentos Operacionais"),
    ("Impostos Federais Pagos",            "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Manutenção Informatica",             "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Taxa Fiscalização",                  "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Material De Uso E Consumo",          "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Outras Despesas",                    "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Internet",                           "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Licenças Ambientais",                "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Seguros",                            "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Convenio Medico",                    "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Despesas Eventuais",                 "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Locação Equipamentos",               "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Consultoria E Assessoria",           "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Serviços De Terceiros",              "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Tarifa De Manutenção Frota",         "ATIVIDADES OPERACIONAIS",   "Despesas"),
    ("Investimento Compra De Equipamentos","ATIVIDADES DE INVESTIMENTO", "Atividades de Investimento"),
]


# ════════════════════════════════════════════════════════════════════════════
# Detector de linha agregadora / cabeçalho / sub-credor
# ════════════════════════════════════════════════════════════════════════════
def _is_aggregator(s):
    s = s.strip()
    if not s: return True
    if s.startswith("%") or s.startswith("(="): return True
    if s.startswith("( =") or s.startswith("( +") or s.startswith("( -"): return True
    up = s.upper()
    if up in {"RECEBIMENTOS OPERACIONAIS", "PAGAMENTOS OPERACIONAIS", "MÚTUOS",
              "BANCOS", "OUTRAS EMPRESAS", "FLUXO DE CAIXA",
              "ENTRADAS", "SAÍDAS", "RECEBIMENTOS", "FORNECEDORES", "DESPESAS",
              "ATIVIDADES DE FINANCIAMENTO", "ATIVIDADES DE INVESTIMENTO",
              "EXTRATO", "PETROS", "POSTOS"}:
        return True
    if s.startswith("Recebimento Acumulado"): return True
    if "(Acumulado)" in s or s.endswith(" Acumulado") or s.endswith(" Acumuladas"): return True
    if s.startswith("DFC -") or s.startswith("Valores em Aberto"): return True
    if s in {"Saldo Inicial", "Saldo Final", "Pagamento Acumulado",
             "Total Recebido - Pagamentos", "NECESSIDADE DE CAIXA (DIA)",
             "GERAÇÃO DE CAIXA ACUMULADO", "Despesas (dia)",
             "Saldo Inicial (Bancário)", "Saldo Final (Bancário)",
             "Total Bancos", "Diferenças Bancos", "Pagamento Autorizado",
             "Diferenças Pag Autorizado", "Pagamentos - Fornecedores (dia)",
             "Pagamentos - Fornecedores (acumulado)",
             "(+) Recebimentos", "(-) Pagamentos",
             "Recebimento Acumulado (exc Ativ Financiam)",
             "% Despesas", "% Fornecedores x Recebimentos",
             # Linhas de controle/conciliação do DFC0X — não são lançamentos reais
             "Recebimentos - Pagamentos", "Recebimentos - Pagamentos (Exceto Aluguel)",
             "Transferencia Bancaria Entre Grupos",
             "Mútuos Recebidos Entre Grupos",
             "Recebimento Supermercados",
             "Mútuos pendentes R", "Transferencia pendente P",
             "Entre Grupos R", "Entre Grupos P",
             "( = ) Diferença", "( = ) Diferença de Recebimentos",
             "( = ) SALDO FINAL (EXC. ALUGUEL)", "( = ) SALDO FINAL CALCULADO"}:
        return True
    return False


# ════════════════════════════════════════════════════════════════════════════
# Extrai plano de contas canônico da aba ANALÍTICO de cada segmento
# ════════════════════════════════════════════════════════════════════════════
def extrair_plano_de_contas(path, sheet_name):
    """Walks rows; detects section headers; returns ordered list
       [(linha, grupo, agrupamento)] do segmento.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    cur_grupo, cur_agrup = None, None
    out = []
    for row in ws.iter_rows(values_only=True):
        if len(row) < 3: continue
        lab = row[2]
        if not lab or not isinstance(lab, str): continue
        s = lab.strip()
        if not s: continue
        up = s.upper()
        if up in SECTION_KEYWORDS:
            cur_grupo, cur_agrup = SECTION_KEYWORDS[up]
            continue
        if _is_aggregator(s): continue
        if cur_grupo is None: continue
        agrup = LINHA_AGRUPAMENTO_OVERRIDE.get(s, cur_agrup)
        out.append((s, cur_grupo, agrup))
    wb.close()
    return out


def normalizar(label, segment, canon_set):
    """Tenta casar `label` com uma linha canônica do segmento. Aplica aliases.
       Retorna a forma canônica OU None se não casar."""
    if not label: return None
    s = label.strip()
    if not s: return None
    if s in canon_set: return s
    a = ALIAS.get(segment, {}).get(s)
    if a and a in canon_set: return a
    return None


# ════════════════════════════════════════════════════════════════════════════
# Parser RESUMO — extrai bloco BANCOS e saldos bancários por dia
# ════════════════════════════════════════════════════════════════════════════
# Mapa exact label → campo do dict `bancos` no doc. Difere por segmento.
RESUMO_BANCOS_FIELDS_POSTOS = {
    "Saldo Inicial (Bancário)":    "saldoInicialBancario",
    "Saldo Final (Bancário)":      "saldoFinalBancario",   # 1º ocorrência
    "Total Bancos":                "totalBancos",
    "Diferenças Bancos":           "diferencasBancos",
    "Pagamento Autorizado":        "pagamentoAutorizado",
    "Diferenças Pag Autorizado":   "diferencasPagAutorizado",
}
# Em Postos o bloco "OUTRAS EMPRESAS" tem 2 linhas após o "Saldo Final (Bancário)"
# principal. Vou identificar a 2ª ocorrência de "Saldo Final (Bancário)" como
# "saldoFinalOutras" via contagem de ocorrências.

RESUMO_BANCOS_FIELDS_OUTRAS = {
    "Saldo Inicial (Bancário)":    "saldoInicialBancario",
    "Saldo Final (Bancário)":      "saldoFinalBancario",
    "Total Bancos":                "totalBancos",
    "Diferenças Bancos":           "diferencasBancos",
    "Pagamento Autorizado":        "pagamentoAutorizado",
    "Diferenças Pag Autorizado":   "diferencasPagAutorizado",
}


def parse_resumo(path, sheet_name, segmento):
    """Lê RESUMO/RESUMO_Mês_Atual: extrai os valores das linhas do bloco
       SALDOS BANCÁRIOS + BANCOS, indexados por dia.
       Retorna (ano, mes, {dia_int: {campo: valor}}).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # Detecta linha das datas
    date_row_idx = None
    for i, r in enumerate(rows[:12]):
        dts = [c for c in r if isinstance(c, dt.datetime)]
        if len(dts) >= 5:
            date_row_idx = i; break
    if date_row_idx is None:
        wb.close()
        raise RuntimeError(f"Datas não encontradas em {sheet_name}")
    date_row = rows[date_row_idx]
    col_to_date = {ci: c.date() for ci, c in enumerate(date_row) if isinstance(c, dt.datetime)}
    dates = list(col_to_date.values())
    ano, mes = dates[0].year, dates[0].month
    fields_map = RESUMO_BANCOS_FIELDS_POSTOS if segmento == "postos" else RESUMO_BANCOS_FIELDS_OUTRAS
    out = defaultdict(dict)  # dia → {campo: valor}
    # Pra Postos: o bloco "OUTRAS EMPRESAS" tem "Saldo Final (Bancário)" e
    # "(-) Pagamentos" — vou diferenciar pelo contador de ocorrência.
    saldo_final_seen = 0
    pagamentos_seen  = 0
    section_outras = False
    section_postos_in_outras = False
    for r in rows[date_row_idx + 1:]:
        if len(r) < 3: continue
        lab = r[2]
        if not lab or not isinstance(lab, str): continue
        s = lab.strip()
        # Detecta entrada no bloco "OUTRAS EMPRESAS" (em Postos) ou "POSTOS" (em Outras)
        if s.upper() == "OUTRAS EMPRESAS" and segmento == "postos":
            section_outras = True
            continue
        if s.upper() == "POSTOS" and segmento == "outras":
            section_postos_in_outras = True
            continue
        if s.upper() == "BANCOS":
            section_outras = False
            section_postos_in_outras = False
            continue
        # Re-mapeia labels dentro do bloco "OUTRAS EMPRESAS"/"POSTOS" pra campos próprios
        if segmento == "postos" and section_outras:
            campo = None
            if s == "Saldo Final (Bancário)": campo = "outrasEmpresasSaldoFinal"
            elif s == "(-) Pagamentos":        campo = "outrasEmpresasPagamentos"
            if campo:
                for ci, date in col_to_date.items():
                    if ci >= len(r): continue
                    v = r[ci]
                    if v is None: continue
                    try: out[date.day][campo] = float(v)
                    except (TypeError, ValueError): pass
            continue
        if segmento == "outras" and section_postos_in_outras:
            campo = None
            if s == "Saldo Final (Bancário)": campo = "postosSaldoFinal"
            elif s == "(-) Pagamentos":        campo = "postosPagamentos"
            if campo:
                for ci, date in col_to_date.items():
                    if ci >= len(r): continue
                    v = r[ci]
                    if v is None: continue
                    try: out[date.day][campo] = float(v)
                    except (TypeError, ValueError): pass
            continue
        # Bloco principal (SALDOS / BANCOS)
        campo = fields_map.get(s)
        if campo is None: continue
        for ci, date in col_to_date.items():
            if ci >= len(r): continue
            v = r[ci]
            if v is None: continue
            try:
                # Não sobrescreve (preserva 1ª ocorrência) — bloco OUTRAS EMPRESAS
                # já foi tratado acima por contexto.
                if campo not in out[date.day]:
                    out[date.day][campo] = float(v)
            except (TypeError, ValueError): pass
    wb.close()
    return ano, mes, dict(out)


# ════════════════════════════════════════════════════════════════════════════
# Parser per-posto diário — REC_POSTO + PG_POSTO + DIN_POSTO (File 2)
# Combina os 3 detalhamentos transacionais em uma estrutura uniforme:
#   {posto: {(ano,mes): {linha: {day: value}}}}
# ════════════════════════════════════════════════════════════════════════════
def _parse_date_br(s):
    """Aceita 'DD/MM/YYYY' string ou datetime; retorna date ou None."""
    if isinstance(s, dt.datetime): return s.date()
    if isinstance(s, dt.date): return s
    if isinstance(s, str):
        try: return dt.datetime.strptime(s.strip(), "%d/%m/%Y").date()
        except Exception: return None
    return None


def _safe_float(v):
    if v is None: return 0.0
    try: return float(v)
    except (TypeError, ValueError): return 0.0


def parse_detalhe_postos(path, canon_postos):
    """Detalhe transacional dos POSTOS — uma linha por lançamento liquidado.
       Combina as 5 abas (BASE_RECEB, REC_POSTO, PAGO, PG_POSTO, DIN_POSTO).
       Retorna {(ano, mes): [items…]} pronto pro detalhamento click."""
    out = defaultdict(list)
    wb = load_workbook(path, read_only=True, data_only=True)

    def _push(ano, mes, dia, linha, posto, valor, nomerazao, observacao, nrotitulo):
        out[(ano, mes)].append({
            "ano": str(ano), "mes": f"{mes:02d}",
            "data": f"{ano:04d}-{mes:02d}-{dia:02d}",
            "linha": linha, "loja": posto, "nroempresa": posto,
            "valor": round(valor, 2),
            "nomerazao": nomerazao or "", "observacao": observacao or "",
            "nrotitulo": nrotitulo or "", "descricao": linha,
        })

    # ── BASE_RECEB (Jan recebimentos)
    ws = wb["BASE_RECEB"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 20: continue
        posto, lab, fornecedor, numero, liq_str, descricao, liquidado = r[2], r[3], r[4], r[5], r[9], r[5], r[19]
        if posto not in POSTOS: continue
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(liq_str)
        if d is None: continue
        v = _safe_float(liquidado)
        if v == 0: continue
        _push(d.year, d.month, d.day, canon, posto, v, fornecedor or "", "", str(numero or ""))

    # ── REC_POSTO (Fev/Mar/Abr)
    ws = wb["REC_POSTO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 15: continue
        posto, lab, credor, numero, pago_val, dia_pgt_str, situacao = r[2], r[3], r[5], r[6], r[13], r[14], r[17] if len(r) > 17 else None
        if posto not in POSTOS: continue
        if situacao and str(situacao).strip().lower() not in ("liquidado", "liquidada", "1"):
            continue
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(dia_pgt_str)
        if d is None: continue
        if d.year == 2026 and d.month == 1: continue   # já em BASE_RECEB
        v = _safe_float(pago_val)
        if v == 0: continue
        _push(d.year, d.month, d.day, canon, posto, v, credor or "", "", str(numero or ""))

    # ── PAGO (Jan pagamentos)
    ws = wb["PAGO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 16: continue
        posto, lab, credor, numero, pgt_str, pago_val, observacao = r[2], r[4], r[6], r[7], r[11], r[15], r[19] if len(r) > 19 else ""
        if posto not in POSTOS: continue
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(pgt_str)
        if d is None: continue
        v = _safe_float(pago_val)
        if v == 0: continue
        _push(d.year, d.month, d.day, canon, posto, -abs(v), credor or "", observacao or "", str(numero or ""))

    # ── PG_POSTO (Mar/Abr pagamentos)
    ws = wb["PG_POSTO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 17: continue
        posto, valores, lab, fornecedor, numero, pag_str, observacao = r[2], r[3], r[4], r[6], r[7], r[11], r[14] if len(r) > 14 else ""
        if posto not in POSTOS: continue
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(pag_str)
        if d is None: continue
        if d.year == 2026 and d.month == 1: continue   # já em PAGO
        v = _safe_float(valores)
        if v == 0: continue
        _push(d.year, d.month, d.day, canon, posto, v, fornecedor or "", observacao or "", str(numero or ""))

    # ── DIN_POSTO (cash em CONTA COFRE)
    DIN_LINHA = "Recebmto de Venda em Dinheiro (Protege)"
    if DIN_LINHA in canon_postos:
        ws = wb["DIN_POSTO"]
        for r in ws.iter_rows(values_only=True, min_row=2):
            if len(r) < 10: continue
            posto, data_str, conta, entrada, saida = r[1], r[2], r[3], r[8], r[9]
            if posto not in POSTOS: continue
            d = _parse_date_br(data_str)
            if d is None: continue
            net = _safe_float(entrada) - _safe_float(saida)
            if net == 0: continue
            _push(d.year, d.month, d.day, DIN_LINHA, posto, net, conta or "", "", "")
    wb.close()
    return dict(out)


def parse_transferencias_postos(path):
    """Lê o arquivo 'transferências bancárias - postos.xlsx' (planilha auxiliar
    fornecida pelo usuário). Schema:
       ano, mês, valores, class, Empresa(razão), Fornecedor, N°titulo, Parc,
       Emissao, Vencimento, Pagamento, Valor
    Retorna {(ano, mes): [item, ...]} no mesmo formato de parse_detalhe_postos.
    A "loja" é o nome curto extraído da razão social (CIDADE OCIDENTAL, EPTG, etc).
    """
    if not os.path.exists(path):
        print(f"   [WARN] {path} não existe — pulando transferências postos.")
        return {}
    out = defaultdict(list)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    n = 0
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r or len(r) < 12: continue
        ano_str, mes_str, _val_neg, lab, empresa, fornecedor, nrotit, parc, emissao, venc, pgt_str, valor = r[:12]
        if not lab or not empresa: continue
        canon = str(lab).strip()    # "Transferencia Bancaria" — já é a linha canônica
        # Posto: extrai o que vier depois de "AUTO POSTO " (ou usa primeira palavra
        # significativa). É só pra rotular o lançamento no drill-down.
        emp = str(empresa).strip()
        upper = emp.upper()
        posto = upper.replace("AUTO POSTO ", "").split(" LTDA")[0].split(" COM")[0].strip()
        if not posto: posto = emp[:20]
        d = _parse_date_br(pgt_str) or _parse_date_br(venc) or _parse_date_br(emissao)
        if d is None: continue
        v = _safe_float(valor)
        if v == 0: continue
        # "valores" na planilha já vem negativo (saída de caixa). Garante sinal.
        v = -abs(v)
        out[(d.year, d.month)].append({
            "ano": str(d.year), "mes": f"{d.month:02d}",
            "data": f"{d.year:04d}-{d.month:02d}-{d.day:02d}",
            "linha": canon, "loja": posto, "nroempresa": posto,
            "valor": round(v, 2),
            "nomerazao": fornecedor or "",
            "observacao": f"{emp} → {fornecedor or ''}".strip(" →"),
            "nrotitulo": str(nrotit or ""), "descricao": canon,
        })
        n += 1
    wb.close()
    print(f"   transferências postos: {n} lançamentos lidos de {path}")
    return dict(out)


def parse_comparativo_protege_postos(path_f1, path_f2, postos_mensal=None):
    """Compara vendas em dinheiro no PDV (DIN_POSTO entradas no cofre) vs
       o que a Protege depositou no banco (BASE_RECEB ou DFC mensal per-posto).
       Retorna {(ano, mes, posto): {vendas, protege}}.
       Fontes:
         vendas_dinheiro = DIN_POSTO.ENTRADA (Jan-Abr 2026, per-posto)
         depositado_protege = BASE_RECEB classe Protege (Dez/25 + Jan/26) +
                              postos_mensal (Jan/Fev/Mar via DFC0X per-posto)
    """
    out = defaultdict(lambda: {"vendas_dinheiro": 0.0, "depositado_protege": 0.0})

    # ── DIN_POSTO (File 2): MêS, POSTO, Data, Conta, ..., ENTRADA, SAIDA
    # ENTRADA = vendas em dinheiro no PDV (entrada no cofre)
    # SAIDA   = retirada Protege (cofre → banco). Não usamos pq Protege deposita
    #           via BASE_RECEB / DFC0X — usar isso pra dep. seria dupla contagem.
    wb2 = load_workbook(path_f2, read_only=True, data_only=True)
    ws = wb2["DIN_POSTO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 10: continue
        posto, data_str, entrada = r[1], r[2], r[8]
        if posto not in POSTOS: continue
        d = _parse_date_br(data_str)
        if d is None: continue
        v = _safe_float(entrada)
        if v == 0: continue
        out[(d.year, d.month, posto)]["vendas_dinheiro"] += v

    # ── BASE_RECEB Protege (File 2): ANO, MÊS, POSTO, DESC, ..., Liquidado(19)
    #     Linhas Protege têm a maioria dos campos None; só ANO/MÊS/POSTO/DESC/Liquidado.
    ws = wb2["BASE_RECEB"]
    PROTEGE_LINES = {
        "Recbto de Venda em Dinheiro (Protege)",
        "Recebmto de Venda em Dinheiro (Protege)",
    }
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 20: continue
        ano, mes, posto, desc, liquidado = r[0], r[1], r[2], r[3], r[19]
        if posto not in POSTOS: continue
        if not desc or str(desc).strip() not in PROTEGE_LINES: continue
        try: ano, mes = int(ano), int(mes)
        except (TypeError, ValueError): continue
        v = _safe_float(liquidado)
        if v == 0: continue
        out[(ano, mes, posto)]["depositado_protege"] += v
    wb2.close()

    # ── postos_mensal (DFC0X Jan/Fev/Mar): linha "Recebmto de Venda em Dinheiro
    # (Protege)" per posto. Só preenche se BASE_RECEB não cobriu aquele (ano,mes,posto).
    if postos_mensal:
        for (ano, mes), p in postos_mensal.items():
            for posto, lvs in p.items():
                v = lvs.get("Recebmto de Venda em Dinheiro (Protege)", 0)
                if v == 0: continue
                k = (ano, mes, posto)
                if out[k]["depositado_protege"] == 0:
                    out[k]["depositado_protege"] = v
    return dict(out)


def parse_taxas_cartoes_postos(path_tarifas):
    """Lê BASE_CARTTX + BANDEIRA e devolve:
         items por (ano, mes, posto, bandeira) com valor bruto, taxa paga,
         taxa negociada (contratual), diferença.
       Adiciona linhas agregadas por bandeira (ano todo)."""
    wb = load_workbook(path_tarifas, read_only=True, data_only=True)

    # BANDEIRA: aba com taxas contratuais
    taxa_contr = {}
    ws = wb["BANDEIRA"]
    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 4: continue
        if r[2] is not None and isinstance(r[3], (int, float)):
            taxa_contr[str(r[2]).strip()] = float(r[3])

    # BASE_CARTTX: ANO, MÊS, EMPRESA, BANDEIRA, Valor, Desconto, Acréscimo, Valor Liquido
    # IMPORTANTE: Tarifa Paga = Valor Bruto − Valor Líquido (Desconto/Acréscimo
    # estão sempre zerados no BASE_CARTTX — esses campos não refletem a tarifa).
    ws = wb["BASE_CARTTX"]
    items = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 8 or r[0] is None: continue
        try:
            ano, mes = int(r[0]), int(r[1])
        except (TypeError, ValueError): continue
        posto = r[2]
        bandeira = str(r[3] or "").strip()
        valor = _safe_float(r[4])
        valor_liq = _safe_float(r[7])
        if posto not in POSTOS or valor == 0: continue
        tarifa_paga = valor - valor_liq
        tx_paga = tarifa_paga / valor if valor else 0
        tx_contr = taxa_contr.get(bandeira, 0)
        tarifa_contr = valor * tx_contr
        diff_rs = tarifa_contr - tarifa_paga
        items.append({
            "ano": ano, "mes": mes, "posto": posto, "bandeira": bandeira,
            "valor_bruto": round(valor, 2),
            "valor_liquido": round(valor_liq, 2),
            "tarifa_paga_rs": round(tarifa_paga, 2),
            "tx_paga_pct": round(tx_paga, 6),
            "tx_contr_pct": round(tx_contr, 6),
            "tarifa_contr_rs": round(tarifa_contr, 2),
            "diff_rs": round(diff_rs, 2),
        })
    wb.close()
    return {"taxa_contratual": taxa_contr, "items": items}


def parse_detalhe_outras(path_f3, canon_outras):
    """Detalhe transacional das OUTRAS — RECEBIMENTOS + PAGAMENTOS (File 3).
       Filtra Pagamentos por col Pagamento preenchida (quitados)."""
    out = defaultdict(list)
    wb = load_workbook(path_f3, read_only=True, data_only=True)

    def _push(ano, mes, dia, linha, empresa, valor, nomerazao, observacao, nrotitulo):
        out[(ano, mes)].append({
            "ano": str(ano), "mes": f"{mes:02d}",
            "data": f"{ano:04d}-{mes:02d}-{dia:02d}",
            "linha": linha, "loja": empresa, "nroempresa": empresa,
            "valor": round(valor, 2),
            "nomerazao": nomerazao or "", "observacao": observacao or "",
            "nrotitulo": nrotitulo or "", "descricao": linha,
        })

    # ── RECEBIMENTOS: Class, Vencimento, Número, Razão Social, CNPJ,
    #    Descrição, Nominal, Total. A coluna Razão Social mudou de "código"
    #    curto (FLUXO/LP/...) pra nome completo ("RETA COMERCIAL LTDA", etc),
    #    então mapeamos por substring antes de filtrar.
    RAZAO_TO_OUTRAS = {
        "FLUXO":  "FLUXO",   # FLUXO CONSULTORIA EMPRESARIAL LTDA
        "LP":     "LP",      # LP´LOGISTICA E GESTAO DE VEICULOS
        "PEGUI":  "PEGUI",   # PEGUI IMOBILIARIA LTDA
        "RETA":   "RETA",    # RETA COMERCIAL LTDA
        "TARES":  "TARES",   # COMERCIAL TARES LTDA
    }
    def _map_outras(raw):
        if not raw: return None
        s = str(raw).upper().strip()
        if s in OUTRAS: return s            # ainda compatível com formato antigo
        for sub, code in RAZAO_TO_OUTRAS.items():
            if sub in s: return code
        return None

    ws = wb["RECEBIMENTOS"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 8: continue
        lab, venc, numero, empresa_raw, descricao, total = r[0], r[1], r[2], r[3], r[5], r[7]
        empresa = _map_outras(empresa_raw)
        if empresa is None: continue
        canon = normalizar(lab, "outras", canon_outras)
        if canon is None: continue
        d = _parse_date_br(venc)
        if d is None: continue
        v = _safe_float(total)
        if v == 0: continue
        _push(d.year, d.month, d.day, canon, empresa, v, descricao or "", "", str(numero or ""))

    # Regra de reclassificação OUTRAS: pagamentos cadastrados como "Transferencia
    # Bancaria" cujo histórico contém EMPRE2 + termos de obra (MANUTEN, REFORM,
    # MATERIAIS, REPAR, CONSTRU) são na verdade gastos com manutenção predial
    # (uso de Vale Empréstimo pra cobrir despesa não-recorrente). Reclassifica
    # pra "Manutenção Predial". Usuário confirmou em 18/05/2026.
    _REGEX_MANUT_PREDIAL = ("MANUTEN", "REFORM", "MATERIAIS", "REPARO", "CONSTRU", "OBRA")
    def _override_class_outras(class_atual, historico):
        c = str(class_atual or "").strip()
        if c != "Transferencia Bancaria": return None
        h = str(historico or "").upper()
        if "EMPRE2" not in h: return None
        if not any(p in h for p in _REGEX_MANUT_PREDIAL): return None
        return "Manutenção Predial"

    # ── PAGAMENTOS: EMPRESA, Class, Lançamento, Estabelecimento, Emissao,
    #    Pagamento, Razão, CNPJ, Espécie, Numero, Vl.Nominal, ..., Vl.Pago,
    #    Historico, Conta Bancaria
    ws = wb["PAGAMENTOS"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 15: continue
        empresa, lab, pag_str, razao, numero, vl_pago, historico = r[0], r[1], r[5], r[6], r[9], r[13], r[14]
        # Override pra Manutenção Predial quando bate a regra
        override = _override_class_outras(lab, historico)
        if override: lab = override
        if empresa not in OUTRAS: continue
        if pag_str is None: continue   # só os quitados (com Pagamento preenchido)
        canon = normalizar(lab, "outras", canon_outras)
        if canon is None: continue
        d = _parse_date_br(pag_str)
        if d is None: continue
        v = _safe_float(vl_pago)
        if v == 0: continue
        _push(d.year, d.month, d.day, canon, empresa, -abs(v), razao or "", str(historico or "").strip(), str(numero or ""))
    wb.close()
    return dict(out)


def parse_postos_diario_detalhe(path, canon_postos):
    """Combina REC_POSTO/PG_POSTO/DIN_POSTO + BASE_RECEB/PAGO (Jan/Fev).
       Retorna {posto: {(ano,mes): {linha: {day_int: val}}}}.
       Linhas fora do plano canônico de Postos são descartadas."""
    out = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
    wb = load_workbook(path, read_only=True, data_only=True)

    def _add(posto, d, canon, valor):
        if posto not in POSTOS: return
        out[posto][(d.year, d.month)][canon][d.day] += valor

    # ── BASE_RECEB (Jan recebimentos — também tem 2025-12 que vamos ignorar)
    # Colunas: ANO,MÊS,POSTO,DESC,FORNECEDOR,Número,Id.,...,Liquidação(9),
    #          Emissão,Vencto,Situação,...,Forma Recebimento,Banco,
    #          Principal(16),Desconto(17),Acrescimo(18),Liquidado(19)
    ws = wb["BASE_RECEB"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 20: continue
        posto, lab, liq_str, liquidado = r[2], r[3], r[9], r[19]
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(liq_str)
        if d is None: continue
        try: v = float(liquidado) if liquidado is not None else 0
        except (TypeError, ValueError): continue
        if v == 0: continue
        _add(posto, d, canon, v)

    # ── REC_POSTO (Jan/Fev/Mar/Abr recebimentos)
    # Vai sobrescrever / somar valores de BASE_RECEB para Jan. Pra evitar
    # duplicação, FILTRA REC_POSTO Jan: só usa se a linha não está em BASE_RECEB.
    # Mais simples: NÃO carrega REC_POSTO Jan; BASE_RECEB é mais detalhado.
    ws = wb["REC_POSTO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 15: continue
        posto, lab, pago_val, dia_pgt_str, situacao = r[2], r[3], r[13], r[14], r[17] if len(r) > 17 else None
        if situacao and str(situacao).strip().lower() not in ("liquidado", "liquidada", "1"):
            continue
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(dia_pgt_str)
        if d is None: continue
        # Pula Jan/2026 (já coberto por BASE_RECEB pra evitar dupla contagem)
        if d.year == 2026 and d.month == 1: continue
        try: v = float(pago_val) if pago_val is not None else 0
        except (TypeError, ValueError): continue
        if v == 0: continue
        _add(posto, d, canon, v)

    # ── PAGO (Jan pagamentos)
    # Colunas: ANO,MÊS,POSTO,ObrigDir,RECEBIMENTO(class!),Empresa,Credor,
    #          Numero,Parc,Emissao,Vencimento,dia pgt(11),Valor,Desconto,
    #          Acrescimo,Pago(15),Conta,...
    ws = wb["PAGO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 16: continue
        posto, lab, pgt_str, pago_val = r[2], r[4], r[11], r[15]
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(pgt_str)
        if d is None: continue
        try: v = float(pago_val) if pago_val is not None else 0
        except (TypeError, ValueError): continue
        if v == 0: continue
        _add(posto, d, canon, -abs(v))   # pagamento → negativo

    # ── PG_POSTO (Mar/Abr pagamentos) — `valores` já vem com sinal negativo
    ws = wb["PG_POSTO"]
    for r in ws.iter_rows(values_only=True, min_row=2):
        if len(r) < 14: continue
        posto, valores, lab, pag_str = r[2], r[3], r[4], r[11]
        canon = normalizar(lab, "postos", canon_postos)
        if canon is None: continue
        d = _parse_date_br(pag_str)
        if d is None: continue
        # Pula Jan/2026 (já coberto por PAGO)
        if d.year == 2026 and d.month == 1: continue
        try: v = float(valores) if valores is not None else 0
        except (TypeError, ValueError): continue
        if v == 0: continue
        _add(posto, d, canon, v)

    # ── DIN_POSTO (vendas em dinheiro / protege)
    DIN_LINHA = "Recebmto de Venda em Dinheiro (Protege)"
    if DIN_LINHA in canon_postos:
        ws = wb["DIN_POSTO"]
        for r in ws.iter_rows(values_only=True, min_row=2):
            if len(r) < 10: continue
            posto, data_str, entrada, saida = r[1], r[2], r[8], r[9]
            d = _parse_date_br(data_str)
            if d is None: continue
            try: ent = float(entrada) if entrada is not None else 0
            except (TypeError, ValueError): ent = 0
            try: sai = float(saida) if saida is not None else 0
            except (TypeError, ValueError): sai = 0
            net = ent - sai
            if net == 0: continue
            _add(posto, d, DIN_LINHA, net)
    wb.close()
    return {posto: {ymk: {l: dict(d) for l, d in lines.items()} for ymk, lines in months.items()}
            for posto, months in out.items()}


def _aplicar_fallback_mensal(detalhe_loja, mensal_loja, tolerancia=1.0):
    """Pra cada linha do mensal, se a soma da detalhe diária diverge
       significativamente (ou está vazia), substitui pelo valor mensal no dia 1.
       Retorna o dict atualizado {linha: {day: val}}.
       Mantém a granularidade diária só pras linhas que batem (±tolerancia)."""
    out = dict(detalhe_loja)  # shallow copy
    for linha, mensal_total in (mensal_loja or {}).items():
        det = out.get(linha, {})
        det_total = sum(det.values())
        if abs(det_total - mensal_total) > tolerancia:
            # Detalhe está incompleto → usa mensal-no-dia-1
            out[linha] = {1: mensal_total}
    return out


# ════════════════════════════════════════════════════════════════════════════
# Parser da aba BANCO (presente em File 1, File 2 e File 3 — mesmo layout).
# Cobre dias de Mar/Abr 2026. Estrutura:
#   col A = data ; col B = TOTAL saldo ; col C = POSTOS saldo ;
#   col D = OUTRAS EMPRESAS saldo ; col F = TOTAL pag aut ;
#   col G = POSTOS pag aut ; col H = OUTRAS EMPRESAS pag aut
# Retorno: { (ano, mes): {dia: {campo: valor}} } pra cada segmento.
# ════════════════════════════════════════════════════════════════════════════
def parse_banco_diario(path, sheet_name):
    """Retorna ({(ano,mes): {dia: {campos…}}}_postos,
                {(ano,mes): {dia: {campos…}}}_outras).
       Campos preenchidos:
         postos:  totalBancos, pagamentoAutorizado,
                  outrasEmpresasSaldoFinal, outrasEmpresasPagamentos
         outras:  totalBancos, pagamentoAutorizado,
                  postosSaldoFinal, postosPagamentos
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    postos = defaultdict(lambda: defaultdict(dict))   # (ano,mes) → dia → campos
    outras = defaultdict(lambda: defaultdict(dict))
    for r in ws.iter_rows(values_only=True):
        if not r or not isinstance(r[0], dt.datetime):
            continue
        d = r[0].date()
        # col B,C,D = saldos; col F,G,H = pagamentos
        try:
            saldo_postos    = float(r[2]) if r[2] is not None else None
            saldo_outras    = float(r[3]) if len(r) > 3 and r[3] is not None else None
            pag_postos      = float(r[6]) if len(r) > 6 and r[6] is not None else None
            pag_outras      = float(r[7]) if len(r) > 7 and r[7] is not None else None
        except (TypeError, ValueError):
            continue
        key = (d.year, d.month)
        if saldo_postos is not None:
            postos[key][d.day]["totalBancos"] = saldo_postos
            outras[key][d.day]["postosSaldoFinal"] = saldo_postos
        if saldo_outras is not None:
            outras[key][d.day]["totalBancos"] = saldo_outras
            postos[key][d.day]["outrasEmpresasSaldoFinal"] = saldo_outras
        if pag_postos is not None:
            postos[key][d.day]["pagamentoAutorizado"] = pag_postos
            outras[key][d.day]["postosPagamentos"] = -abs(pag_postos)
        if pag_outras is not None:
            outras[key][d.day]["pagamentoAutorizado"] = pag_outras
            postos[key][d.day]["outrasEmpresasPagamentos"] = -abs(pag_outras)
    wb.close()
    # Converte pra dict normal
    return ({k: dict(v) for k, v in postos.items()},
            {k: dict(v) for k, v in outras.items()})


# ════════════════════════════════════════════════════════════════════════════
# Parser File 1/3: ANALÍTICO diário (col C = linha, col D+ = dias)
# ════════════════════════════════════════════════════════════════════════════
def parse_analitico_diario(path, sheet_name, segment, canon_set):
    """Retorna (ano, mes, {linha_canonica: {dia: valor}})."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # Detecta linha das datas (>=5 datetimes)
    date_row_idx = None
    for i, r in enumerate(rows[:12]):
        dts = [c for c in r if isinstance(c, dt.datetime)]
        if len(dts) >= 5:
            date_row_idx = i
            break
    if date_row_idx is None:
        wb.close()
        raise RuntimeError(f"Datas não encontradas em {sheet_name}")
    date_row = rows[date_row_idx]
    col_to_date = {ci: c.date() for ci, c in enumerate(date_row) if isinstance(c, dt.datetime)}
    dates = list(col_to_date.values())
    ano, mes = dates[0].year, dates[0].month
    per_linha = defaultdict(lambda: defaultdict(float))
    for r in rows[date_row_idx + 1:]:
        if len(r) < 3: continue
        lab = r[2]
        if not lab or not isinstance(lab, str): continue
        canon = normalizar(lab, segment, canon_set)
        if canon is None: continue
        for ci, date in col_to_date.items():
            if ci >= len(r): continue
            v = r[ci]
            if v is None or v == 0: continue
            try: v = float(v)
            except (TypeError, ValueError): continue
            per_linha[canon][date.day] += v
    wb.close()
    return ano, mes, dict(per_linha)


# ════════════════════════════════════════════════════════════════════════════
# Parser File 2: DFC0X (mensal por loja)
# ════════════════════════════════════════════════════════════════════════════
def parse_dfc_mensal_por_loja(path, sheet_name, canon_postos, canon_outras):
    """Lê DFC0X — cols D-N = P01..P11, cols Q-U = FLUXO/LP/PEGUI/RETA/TARES.
       Aplica plano-de-contas POR SEGMENTO (postos × outras).
       Retorna (ano, mes, {posto: {linha: val}}, {empresa: {linha: val}})."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    date_cell = rows[1][2]
    ano, mes = (date_cell.year, date_cell.month) if isinstance(date_cell, dt.datetime) else (None, None)
    header = rows[2]
    col_loja = {}
    for ci, c in enumerate(header):
        if not c: continue
        s = str(c).strip()
        if s in POSTOS or s in OUTRAS:
            col_loja[ci] = s
    postos = defaultdict(lambda: defaultdict(float))
    outras = defaultdict(lambda: defaultdict(float))
    for r in rows[3:]:
        if len(r) < 3: continue
        lab = r[2]
        if not lab or not isinstance(lab, str): continue
        for ci, loja in col_loja.items():
            if ci >= len(r): continue
            v = r[ci]
            if v is None or v == 0: continue
            try: v = float(v)
            except (TypeError, ValueError): continue
            seg = "postos" if loja in POSTOS else "outras"
            canon_set = canon_postos if seg == "postos" else canon_outras
            canon = normalizar(lab, seg, canon_set)
            if canon is None: continue
            if seg == "postos":
                postos[loja][canon] += v
            else:
                outras[loja][canon] += v
    wb.close()
    return ano, mes, dict(postos), dict(outras)


# ════════════════════════════════════════════════════════════════════════════
# Gerador de doc mensal
# ════════════════════════════════════════════════════════════════════════════
def montar_doc_mensal(ano, mes, valores_por_linha_e_dia, *, segmento, taxonomia, agrupamentos, loja=None, bancos=None):
    """taxonomia: lista ordenada [(linha, grupo, agrupamento)] do segmento.
       bancos: opcional, dict {dia_int: {campo: valor}} extraído do RESUMO.
    """
    dias_mes = calendar.monthrange(ano, mes)[1]
    dias = [f"{d:02d}" for d in range(1, dias_mes + 1)]
    linhas = [l for l, _, _ in taxonomia]
    linha_to_idx = {l: i for i, l in enumerate(linhas)}
    grupo_to_idx = {g: i for i, g in enumerate(GRUPOS)}
    agrup_to_idx = {a: i for i, a in enumerate(agrupamentos)}

    porGrupo_acc = defaultdict(float)
    porAgrup_acc = defaultdict(float)
    porLinha_arr = []
    for linha, grupo, agrup in taxonomia:
        n_idx = linha_to_idx[linha]
        g_idx = grupo_to_idx[grupo]
        a_idx = agrup_to_idx[agrup]
        vals = valores_por_linha_e_dia.get(linha, {})
        for d_int, v in vals.items():
            if v == 0: continue
            d_idx = d_int - 1
            porLinha_arr.append({"d": d_idx, "g": g_idx, "a": a_idx, "n": n_idx, "v": round(v, 2)})
            porAgrup_acc[(d_idx, g_idx, a_idx)] += v
            porGrupo_acc[(d_idx, g_idx)] += v
    porAgrupamento = [{"d": d, "g": g, "a": a, "v": round(v, 2)}
                      for (d, g, a), v in porAgrup_acc.items() if abs(v) >= 0.005]
    porGrupo = [{"d": d, "g": g, "v": round(v, 2)}
                for (d, g), v in porGrupo_acc.items() if abs(v) >= 0.005]
    # Bloco BANCOS (RESUMO): {dia_str: {campo: valor}} — só pros meses que
    # têm RESUMO disponível (postos Abr/Mai; outras Abr).
    bancos_out = None
    if bancos:
        bancos_out = {f"{d:02d}": {k: round(v, 2) for k, v in fields.items()}
                      for d, fields in bancos.items() if fields}
    doc = {
        "ano": ano, "mes": mes, "v": 2,
        "segmento": segmento, "loja": loja or "",
        "dim": {"dias": dias, "grupos": GRUPOS, "agrupamentos": agrupamentos, "linhas": linhas},
        "porLinha": porLinha_arr,
        "porAgrupamento": porAgrupamento,
        "porGrupo": porGrupo,
    }
    if bancos_out:
        doc["bancos"] = bancos_out
    return doc


# ════════════════════════════════════════════════════════════════════════════
# Pipeline principal
# ════════════════════════════════════════════════════════════════════════════
def main():
    os.makedirs(OUT_POSTOS, exist_ok=True)
    os.makedirs(OUT_OUTRAS, exist_ok=True)
    # Limpa SÓ os arquivos gerados por esse ETL (mensais + detalhe + meta).
    # NÃO apaga arquivos de outros scripts (saldos_iniciais.json,
    # comparativo_protege.json, taxas_cartoes.json).
    PROTECTED = {"saldos_iniciais.json", "comparativo_protege.json", "taxas_cartoes.json"}
    for folder in (OUT_POSTOS, OUT_OUTRAS):
        for f in os.listdir(folder):
            if not f.endswith(".json"): continue
            if f in PROTECTED: continue
            os.remove(os.path.join(folder, f))

    print("→ Extraindo planos de contas canônicos…")
    taxonomia_postos = extrair_plano_de_contas(F1, "ANALÍTICO_Mês_Atual")
    taxonomia_outras = extrair_plano_de_contas(F3, "ANALÍTICO")
    # Estende canonical de outras com linhas que aparecem no DFC0X de
    # PEGUI/TARES (que vendem combustível/produtos) mas não estão no
    # canonical original do File 3 ANALÍTICO.
    canon_outras_set = {l for l, _, _ in taxonomia_outras}
    for tup in LINHAS_OUTRAS_EXTRAS:
        if tup[0] not in canon_outras_set:
            taxonomia_outras.append(tup)
            canon_outras_set.add(tup[0])
    canon_postos = {l for l, _, _ in taxonomia_postos}
    canon_outras = {l for l, _, _ in taxonomia_outras}
    print(f"   POSTOS: {len(taxonomia_postos)} linhas")
    print(f"   OUTRAS: {len(taxonomia_outras)} linhas")

    # Agrupamentos efetivamente usados por cada segmento (preservando ordem)
    def _agrup_seg(tax):
        seen, out = set(), []
        for _, _, a in tax:
            if a not in seen:
                seen.add(a); out.append(a)
        return out
    agrups_postos = _agrup_seg(taxonomia_postos)
    agrups_outras = _agrup_seg(taxonomia_outras)
    print(f"   POSTOS agrupamentos: {agrups_postos}")
    print(f"   OUTRAS agrupamentos: {agrups_outras}")

    # ──────────────────────────────────────────────────────────────────────
    # File 2 — Jan/Fev/Mar mensal por loja
    # ──────────────────────────────────────────────────────────────────────
    print("→ Lendo File 2 (DFC mensal Jan/Fev/Mar)…")
    postos_mensal, outras_mensal = {}, {}
    for sn in ["DFC01", "DFC02", "DFC03"]:
        ano, mes, p, o = parse_dfc_mensal_por_loja(F2, sn, canon_postos, canon_outras)
        postos_mensal[(ano, mes)] = p
        outras_mensal[(ano, mes)] = o
        print(f"   {sn}: ano={ano} mes={mes}  postos={list(p)}  outras={list(o)}")

    print("→ Lendo aba BANCO (saldos+pag.autorizados diários, Mar/Abr)…")
    postos_bancos_diario, outras_bancos_diario = parse_banco_diario(F2, "BANCO")
    for k, v in postos_bancos_diario.items():
        print(f"   POSTOS bancos {k}: {len(v)} dias")
    for k, v in outras_bancos_diario.items():
        print(f"   OUTRAS bancos {k}: {len(v)} dias")

    print("→ Lendo Comparativo Protege Postos (DIN_POSTO + BASE_RECEB + DFC0X)…")
    protege_postos = parse_comparativo_protege_postos(F1, F2, postos_mensal)
    print(f"   {len(protege_postos)} (ano,mes,posto) entries")

    print("→ Lendo Taxas Cartões Postos (BASE_CARTTX + BANDEIRA)…")
    F_TARIFAS = "/mnt/controller/03 - POSTOS/01 - Demonstrativos/Resultado/04 - Tarifas Cartões Por Bandeira e TEF Fixo e Móvel.xlsx"
    try:
        taxas_postos = parse_taxas_cartoes_postos(F_TARIFAS)
        print(f"   {len(taxas_postos['items'])} items · {len(taxas_postos['taxa_contratual'])} bandeiras")
    except FileNotFoundError:
        print(f"   [WARN] arquivo de tarifas não encontrado em {F_TARIFAS}")
        taxas_postos = {"taxa_contratual": {}, "items": []}

    print("→ Lendo detalhe per-posto diário (REC_POSTO + PG_POSTO + DIN_POSTO)…")
    postos_detalhe = parse_postos_diario_detalhe(F2, canon_postos)
    print("→ Lendo detalhe transacional Postos (5 abas) e Outras (RECEB+PAG)…")
    postos_detalhe_items = parse_detalhe_postos(F2, canon_postos)
    outras_detalhe_items = parse_detalhe_outras(F3, canon_outras)
    # Mescla transferências bancárias dos postos (arquivo auxiliar fornecido
    # pelo usuário em 18/05/2026 — preenche o drill-down de Transferencia Bancaria
    # com Empresa/Fornecedor/Nº título individualizados).
    transf_postos = parse_transferencias_postos(F4_TRANSF_POSTOS)
    for chave, items in transf_postos.items():
        postos_detalhe_items.setdefault(chave, []).extend(items)
    for (ano, mes), items in sorted(postos_detalhe_items.items()):
        print(f"   POSTOS detalhe {ano}-{mes:02d}: {len(items)} lançamentos")
    for (ano, mes), items in sorted(outras_detalhe_items.items()):
        print(f"   OUTRAS detalhe {ano}-{mes:02d}: {len(items)} lançamentos")
    # Sumário pra debug
    for posto in sorted(postos_detalhe):
        meses_p_set = sorted(postos_detalhe[posto].keys())
        total_entries = sum(sum(len(d) for d in lines.values())
                            for lines in postos_detalhe[posto].values())
        print(f"   {posto}: {len(meses_p_set)} meses, {total_entries} entries (linha-dia)")

    # ──────────────────────────────────────────────────────────────────────
    # File 1 / File 3 — diário consolidado Apr/Mai
    # ──────────────────────────────────────────────────────────────────────
    print("→ Lendo File 1 (ANALÍTICO diário Postos Apr/Mai)…")
    postos_diario = {}
    postos_bancos_resumo = {}
    for sn_an, sn_re in [("ANALÍTICO_Mês_Atual",   "RESUMO_Mês_Atual"),
                         ("ANALÍTICO_Próximo_Mês", "RESUMO_Próximo_Mês")]:
        ano, mes, d = parse_analitico_diario(F1, sn_an, "postos", canon_postos)
        postos_diario[(ano, mes)] = d
        _, _, b = parse_resumo(F1, sn_re, "postos")
        postos_bancos_resumo[(ano, mes)] = b
        print(f"   {sn_an}: ano={ano} mes={mes} #linhas={len(d)} #dias-bancos={len(b)}")

    print("→ Lendo File 3 (ANALÍTICO diário Outras)…")
    ano, mes, d3 = parse_analitico_diario(F3, "ANALÍTICO", "outras", canon_outras)
    outras_diario = {(ano, mes): d3}
    _, _, b3 = parse_resumo(F3, "RESUMO", "outras")
    outras_bancos_resumo = {(ano, mes): b3}
    print(f"   ANALÍTICO: ano={ano} mes={mes} #linhas={len(d3)} #dias-bancos={len(b3)}")

    # Merge bancos: BANCO (Mar/Abr) + RESUMO (Abr/Mai postos; Abr outras).
    # RESUMO tem mais campos → tem prioridade nos overlaps.
    def _merge_bancos(banco, resumo):
        """{dia: {campo: val}} ∪ {dia: {campo: val}}. RESUMO sobrescreve."""
        out = {}
        for d, fields in (banco or {}).items():
            out.setdefault(d, {}).update(fields)
        for d, fields in (resumo or {}).items():
            out.setdefault(d, {}).update(fields)
        return out
    def _bancos_for(seg, ano, mes):
        if seg == "postos":
            return _merge_bancos(postos_bancos_diario.get((ano, mes)),
                                 postos_bancos_resumo.get((ano, mes)))
        return _merge_bancos(outras_bancos_diario.get((ano, mes)),
                             outras_bancos_resumo.get((ano, mes)))

    # ──────────────────────────────────────────────────────────────────────
    # POSTOS — gera docs
    #   Estratégia per-loja: usa postos_detalhe (REC_POSTO/PG_POSTO/DIN_POSTO)
    #   pra granularidade DIÁRIA real. Fallback no postos_mensal (DFC0X
    #   monthly per-posto) só pros meses/postos sem detalhe.
    # ──────────────────────────────────────────────────────────────────────
    print("→ Gerando docs POSTOS…")
    anos_p, meses_p = set(), set()
    # Conjunto de (ano,mes) que aparecem em postos_mensal OU postos_detalhe
    todos_meses_postos = set(postos_mensal.keys())
    for posto, mds in postos_detalhe.items():
        todos_meses_postos.update(mds.keys())
    todos_meses_postos.update(postos_diario.keys())

    for (ano, mes) in sorted(todos_meses_postos):
        anos_p.add(ano); meses_p.add(mes)
        consol = defaultdict(lambda: defaultdict(float))   # linha → {day: val}
        # Por loja: combina detalhe diário com fallback mensal por linha
        for posto in POSTOS:
            detalhe = postos_detalhe.get(posto, {}).get((ano, mes), {})
            mensal  = postos_mensal.get((ano, mes), {}).get(posto, {})
            if not detalhe and not mensal: continue
            # Pra cada linha do mensal: se a soma do detalhe não bate, força o
            # valor mensal no dia 1 (preserve total correto). Linhas só no
            # detalhe ficam com a distribuição diária.
            v_loja = _aplicar_fallback_mensal(detalhe, mensal)
            doc = montar_doc_mensal(ano, mes, v_loja, segmento="postos",
                                    taxonomia=taxonomia_postos,
                                    agrupamentos=agrups_postos, loja=posto)
            with open(f"{OUT_POSTOS}/{ano}-{mes:02d}__{posto}.json", "w", encoding="utf-8") as f:
                json.dump(doc, f, ensure_ascii=False)
            for l, dmap in v_loja.items():
                for day, v in dmap.items():
                    consol[l][day] += v
        # Consolidado: prefere ANALÍTICO (postos_diario) quando disponível
        # (é a fonte canônica do RESUMO/ANALÍTICO); senão usa a soma das lojas.
        if (ano, mes) in postos_diario:
            v_cons = postos_diario[(ano, mes)]
        else:
            v_cons = {l: dict(d) for l, d in consol.items()}
        doc = montar_doc_mensal(ano, mes, v_cons, segmento="postos",
                                taxonomia=taxonomia_postos,
                                agrupamentos=agrups_postos,
                                bancos=_bancos_for("postos", ano, mes))
        with open(f"{OUT_POSTOS}/{ano}-{mes:02d}.json", "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False)

    # Detalhe transacional por (ano, mes) — alimenta o drill-down (click numa célula)
    for (ano, mes), items in postos_detalhe_items.items():
        with open(f"{OUT_POSTOS}/detalhe_{ano}-{mes:02d}.json", "w", encoding="utf-8") as f:
            json.dump({"items": items}, f, ensure_ascii=False)

    # ── Comparativo Protege (formato amigável pro front)
    protege_items = []
    for (ano, mes, posto), v in sorted(protege_postos.items()):
        vendas = round(v["vendas_dinheiro"], 2)
        protege = round(v["depositado_protege"], 2)
        diff = round(vendas - protege, 2)
        protege_items.append({
            "ano": ano, "mes": mes, "posto": posto,
            "vendas_dinheiro": vendas,
            "depositado_protege": protege,
            "diferenca": diff,
            "diferenca_pct": round((diff / vendas) if vendas else 0, 6),
        })
    with open(f"{OUT_POSTOS}/comparativo_protege.json", "w", encoding="utf-8") as f:
        json.dump({"items": protege_items, "geradoEm": dt.datetime.now().isoformat()},
                  f, ensure_ascii=False)

    # ── Taxas Cartões
    with open(f"{OUT_POSTOS}/taxas_cartoes.json", "w", encoding="utf-8") as f:
        json.dump({**taxas_postos, "geradoEm": dt.datetime.now().isoformat()},
                  f, ensure_ascii=False)

    meta_postos = {
        "geradoEm": dt.datetime.now().isoformat(),
        "segmento": "postos",
        "dimensoes": {
            "anos":  sorted(anos_p),
            "meses": sorted(meses_p),
            "lojas": POSTOS,
            "grupos": GRUPOS,
            "agrupamentos": agrups_postos,
            "linhas": [l for l, _, _ in taxonomia_postos],
        },
        # Hierarquia explícita: usada pelo dashboard pra montar a árvore
        # de linhas no DFC Diário/Anual quando segmento != supermercados.
        "taxonomia": [{"nome": l, "grupo": g, "agrupamento": a}
                      for l, g, a in taxonomia_postos],
        "obs": "Jan/Fev/Mar: mensal por posto (lançado no dia 1). Abr/Mai: diário consolidado.",
    }
    with open(f"{OUT_POSTOS}/meta.json", "w", encoding="utf-8") as f:
        json.dump(meta_postos, f, ensure_ascii=False, indent=2)

    # ──────────────────────────────────────────────────────────────────────
    # OUTRAS — gera docs
    # ──────────────────────────────────────────────────────────────────────
    print("→ Gerando docs OUTRAS…")
    anos_o, meses_o = set(), set()
    for (ano, mes), o in outras_mensal.items():
        anos_o.add(ano); meses_o.add(mes)
        consol = defaultdict(lambda: defaultdict(float))
        for emp, lvs in o.items():
            for l, v in lvs.items(): consol[l][1] += v
            v_loja = {l: {1: v} for l, v in lvs.items()}
            doc = montar_doc_mensal(ano, mes, v_loja, segmento="outras",
                                    taxonomia=taxonomia_outras,
                                    agrupamentos=agrups_outras, loja=emp)
            with open(f"{OUT_OUTRAS}/{ano}-{mes:02d}__{emp}.json", "w", encoding="utf-8") as f:
                json.dump(doc, f, ensure_ascii=False)
        v_cons = {l: dict(d) for l, d in consol.items()}
        doc = montar_doc_mensal(ano, mes, v_cons, segmento="outras",
                                taxonomia=taxonomia_outras,
                                agrupamentos=agrups_outras,
                                bancos=_bancos_for("outras", ano, mes))
        with open(f"{OUT_OUTRAS}/{ano}-{mes:02d}.json", "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False)
    for (ano, mes), lvs in outras_diario.items():
        anos_o.add(ano); meses_o.add(mes)
        doc = montar_doc_mensal(ano, mes, lvs, segmento="outras",
                                taxonomia=taxonomia_outras,
                                agrupamentos=agrups_outras,
                                bancos=_bancos_for("outras", ano, mes))
        with open(f"{OUT_OUTRAS}/{ano}-{mes:02d}.json", "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False)

    # Detalhe transacional Outras (RECEBIMENTOS + PAGAMENTOS quitados)
    for (ano, mes), items in outras_detalhe_items.items():
        with open(f"{OUT_OUTRAS}/detalhe_{ano}-{mes:02d}.json", "w", encoding="utf-8") as f:
            json.dump({"items": items}, f, ensure_ascii=False)

    meta_outras = {
        "geradoEm": dt.datetime.now().isoformat(),
        "segmento": "outras",
        "dimensoes": {
            "anos":  sorted(anos_o),
            "meses": sorted(meses_o),
            "lojas": OUTRAS,
            "grupos": GRUPOS,
            "agrupamentos": agrups_outras,
            "linhas": [l for l, _, _ in taxonomia_outras],
        },
        "taxonomia": [{"nome": l, "grupo": g, "agrupamento": a}
                      for l, g, a in taxonomia_outras],
        "obs": "Jan/Fev/Mar: mensal por empresa (lançado no dia 1). Abr: diário consolidado.",
    }
    with open(f"{OUT_OUTRAS}/meta.json", "w", encoding="utf-8") as f:
        json.dump(meta_outras, f, ensure_ascii=False, indent=2)

    print()
    print(f"✓ Postos:  {sum(1 for _ in os.scandir(OUT_POSTOS) if _.is_file())} arquivos")
    print(f"✓ Outras:  {sum(1 for _ in os.scandir(OUT_OUTRAS) if _.is_file())} arquivos")

    # Override do "Pagamento Autorizado" (DFC Consolidado postos/outras) a partir
    # da PLANILHA DE CONCILIAÇÃO (linha AUTORIZADO DIRETORIA, aba PAINEL). Essa é
    # a fonte autoritativa do pagamento autorizado do dia por posto/empresa —
    # corrige os erros do RESUMO (que conflundia postos×outras em alguns dias).
    # Falha graciosamente se o share \\10.61.1.13\cvl não estiver acessível.
    try:
        import gera_pagamento_autorizado as gpa
        base = gpa.stage_local()
        if base:
            somas, _audit, _dias = gpa.coleta(base)
            rel = gpa.aplica(somas, _dias)
            print(f"✓ Pagamento Autorizado (conciliação): {len(rel)} dias sobrescritos")
        else:
            print("· Conciliação indisponível — pagamentoAutorizado mantido do RESUMO")
    except Exception as e:
        print(f"! Falha ao aplicar conciliação (pagamentoAutorizado): {e}")


if __name__ == "__main__":
    main()
