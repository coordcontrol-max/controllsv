"""Gera planilha modelo preenchida com os dados da conta Equatorial Goiás
(UC 2.960.249.012-88 · ref ABR/2026 · ANTONIO BATISTA LIMA SILVA).

Segue a estrutura de colunas da BASE do `Análise Energia.xlsx` usada pelo
`etl_energia.py`, e adiciona aba "Detalhes" com todos os campos da fatura.
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("/root/projeto_dre/modelo_conta_energia_ABR2026.xlsx")

# ---------- Dados da fatura ----------
UC = "2.960.249.012-88"
CLIENTE = "ANTONIO BATISTA LIMA SILVA"
CPF = "821.775.621-04"
ENDERECO = "RUA SEM NOME, Q. 27, L. 18, S/N · JARDIM ORIENTE · VALPARAISO DE GOIAS · GO"
CLASSIFICACAO = "B1 RESIDENCIAL - RESIDENCIAL NORMAL CONVENCIONAL"
FORNECIMENTO = "TRIFÁSICO"
DISTRIBUIDORA = "Equatorial Goiás Distribuidora de Energia S.A."
CNPJ_DIST = "01.543.032/0001-04"

REF = "ABR/2026"
ANO = 2026
MES_NUM = 4
MES_LBL = "ABR"

VENC = date(2026, 5, 12)
LEITURA_ANT = date(2026, 3, 30)
LEITURA_ATUAL = date(2026, 4, 29)
PROX_LEITURA = date(2026, 5, 29)
EMISSAO = "29/04/2026 11:46:44"
QTD_DIAS = 30

CONSUMO_KWH = 1144
PRECO_UNIT = 1.125925         # com tributos
VALOR_CONSUMO = 1288.06        # FORNECIMENTO CONSUMO kWh
CIP = 60.21                    # Contrib. Ilum. Pública
TOTAL = 1348.27

# Tributos
ICMS_BASE = 1288.06
ICMS_ALIQ = 0.19
ICMS_VAL  = 244.73
COFINS_BASE = 1043.33
COFINS_ALIQ = 0.01822
COFINS_VAL  = 19.01
PIS_BASE  = 1043.33
PIS_ALIQ  = 0.003918
PIS_VAL   = 4.09

# Medidor
MEDIDOR = "13985177-1"
LEIT_ANT = 2288
LEIT_ATUAL = 3432
CONSTANTE = 1.0

# NFe
NFE_NUMERO = 189059964
NFE_SERIE = 0
CFOP = "5258 — Venda de energia elétrica para não contribuinte"
CHAVE = "52260401543032000104660001890599642000572762"
PROTOCOLO = "3522600014083717 — 29/04/2026 13:20:48"

# Histórico de consumo (kWh / dias)
HISTORICO = [
    ("ABR/26", 1144.0, 30, "LIDA"),
    ("MAR/26", 1193.0, 31, "LIDA"),
    ("FEV/26", 0.0, 29, ""),
    ("JAN/26", 0.0, 30, ""),
    ("DEZ/25", 0.0, 31, ""),
    ("NOV/25", 0.0, 31, ""),
    ("OUT/25", 0.0, 32, ""),
    ("SET/25", 0.0, 30, ""),
    ("AGO/25", 0.0, 30, ""),
    ("JUL/25", 0.0, 31, ""),
    ("JUN/25", 0.0, 29, ""),
    ("MAI/25", 0.0, 31, ""),
    ("ABR/25", 0.0, 31, ""),
]
MEDIA = 179.77

# ---------- Estilos ----------
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2F855A")
LBL_FONT = Font(bold=True)
LBL_FILL = PatternFill("solid", fgColor="E6F4EA")
TITLE_FONT = Font(bold=True, size=14, color="1A3A2F")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

BRL = 'R$ #,##0.00'
PCT = '0.0000%'
KWH = '#,##0" kWh"'
DATE = 'dd/mm/yyyy'


def auto_width(ws, extra=2):
    for col_cells in ws.columns:
        col = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col].width = min(max_len + extra, 50)


# ---------- Workbook ----------
wb = Workbook()

# ============ Aba 1: BASE (formato compatível com etl_energia.py) ============
ws = wb.active
ws.title = "BASE"

COLS = [
    "Loja", "ANO", "MESES", "Mês", "Nº Inscrição",
    "Valor Cativo R$", "Dem", "Ult Dem", "Multas e Juros", "Taxas Diversas",
    "Valor s/ Taxas", "Valor c/ Injeção", "Valor GD/ML", "Desconto", "Desc (%)",
    "Total", "Venda", "(%) S/ Venda", "Venc Fornec", "Qtd dias",
    "Leitura", "Venc GD/ML", "TOTAL KWh", "INJEÇÃO", "kWh FORNEC R$",
    "kWh GD/ML R$", "FORNEC + GD/ML (R$/kWh)", "(%) SOBRE INJEÇÃO", "kWh Dia",
]

for i, name in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=i, value=name)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

row = [
    CLIENTE, ANO, MES_LBL, MES_NUM, UC,
    TOTAL, None, None, 0.0, CIP,
    VALOR_CONSUMO, None, 0.0, 0.0, 0.0,
    TOTAL, None, None, VENC, QTD_DIAS,
    LEITURA_ATUAL, None, CONSUMO_KWH, 0.0, PRECO_UNIT,
    0.0, PRECO_UNIT, 0.0, round(CONSUMO_KWH / QTD_DIAS, 2),
]
for i, v in enumerate(row, start=1):
    c = ws.cell(row=2, column=i, value=v)
    c.border = BORDER
    c.alignment = CENTER

# formatos
brl_cols = [6, 9, 10, 11, 12, 13, 14, 16, 17, 25, 26, 27]
for col in brl_cols:
    ws.cell(row=2, column=col).number_format = BRL
ws.cell(row=2, column=15).number_format = PCT
ws.cell(row=2, column=18).number_format = PCT
ws.cell(row=2, column=19).number_format = DATE
ws.cell(row=2, column=21).number_format = DATE
ws.cell(row=2, column=22).number_format = DATE
ws.cell(row=2, column=23).number_format = KWH
ws.cell(row=2, column=24).number_format = KWH
ws.cell(row=2, column=29).number_format = KWH

auto_width(ws)
ws.freeze_panes = "A2"

# ============ Aba 2: Detalhes da Fatura ============
ws2 = wb.create_sheet("Detalhes")

ws2.merge_cells("A1:D1")
t = ws2["A1"]
t.value = "FATURA EQUATORIAL GOIÁS · ABR/2026"
t.font = TITLE_FONT
t.alignment = CENTER
ws2.row_dimensions[1].height = 26

def pair(row, label, value, fmt=None):
    a = ws2.cell(row=row, column=1, value=label)
    b = ws2.cell(row=row, column=2, value=value)
    a.font = LBL_FONT
    a.fill = LBL_FILL
    a.alignment = LEFT
    a.border = BORDER
    b.border = BORDER
    b.alignment = LEFT
    if fmt:
        b.number_format = fmt

# Bloco identificação
ws2["A3"] = "IDENTIFICAÇÃO"
ws2["A3"].font = HDR_FONT
ws2["A3"].fill = HDR_FILL
ws2.merge_cells("A3:D3")
ws2["A3"].alignment = CENTER

pair(4,  "Distribuidora",       DISTRIBUIDORA)
pair(5,  "CNPJ Distribuidora",  CNPJ_DIST)
pair(6,  "Cliente",             CLIENTE)
pair(7,  "CPF",                 CPF)
pair(8,  "Endereço",            ENDERECO)
pair(9,  "Unidade Consumidora", UC)
pair(10, "Classificação",       CLASSIFICACAO)
pair(11, "Tipo Fornecimento",   FORNECIMENTO)
pair(12, "Tensão Nominal",      "220 V (min 200,2 V · max 231,0 V)")

# Bloco período / leitura
ws2["A14"] = "PERÍODO E LEITURA"
ws2["A14"].font = HDR_FONT
ws2["A14"].fill = HDR_FILL
ws2.merge_cells("A14:D14")
ws2["A14"].alignment = CENTER

pair(15, "Referência",          REF)
pair(16, "Vencimento",          VENC, DATE)
pair(17, "Leitura Anterior",    LEITURA_ANT, DATE)
pair(18, "Leitura Atual",       LEITURA_ATUAL, DATE)
pair(19, "Nº de dias",          QTD_DIAS)
pair(20, "Próxima leitura",     PROX_LEITURA, DATE)
pair(21, "Emissão NFe",         EMISSAO)

# Bloco consumo / medidor
ws2["A23"] = "CONSUMO E MEDIDOR"
ws2["A23"].font = HDR_FONT
ws2["A23"].fill = HDR_FILL
ws2.merge_cells("A23:D23")
ws2["A23"].alignment = CENTER

pair(24, "Medidor",                 MEDIDOR)
pair(25, "Leitura anterior (kWh)",  LEIT_ANT)
pair(26, "Leitura atual (kWh)",     LEIT_ATUAL)
pair(27, "Constante",               CONSTANTE)
pair(28, "Consumo (kWh)",           CONSUMO_KWH, KWH)
pair(29, "Preço unit. c/ tributos", PRECO_UNIT, 'R$ 0.000000')
pair(30, "Tarifa unit. (R$/kWh)",   0.891810, 'R$ 0.000000')
pair(31, "Consumo médio diário",    round(CONSUMO_KWH/QTD_DIAS, 2), KWH)

# Bloco valores
ws2["A33"] = "COMPOSIÇÃO DA FATURA"
ws2["A33"].font = HDR_FONT
ws2["A33"].fill = HDR_FILL
ws2.merge_cells("A33:D33")
ws2["A33"].alignment = CENTER

pair(34, "Fornecimento — Consumo kWh", VALOR_CONSUMO, BRL)
pair(35, "Contrib. Ilum. Pública (CIP)", CIP, BRL)
pair(36, "Multas e juros (mês atual)", 0.0, BRL)
pair(37, "TOTAL A PAGAR",              TOTAL, BRL)
ws2["A37"].font = Font(bold=True)
ws2["B37"].font = Font(bold=True, color="C0392B")

# Bloco tributos (mini-tabela)
ws2["A39"] = "TRIBUTOS"
ws2["A39"].font = HDR_FONT
ws2["A39"].fill = HDR_FILL
ws2.merge_cells("A39:D39")
ws2["A39"].alignment = CENTER

hdr = ["Tributo", "Base (R$)", "Alíquota", "Valor (R$)"]
for i, h in enumerate(hdr, start=1):
    c = ws2.cell(row=40, column=i, value=h)
    c.font = LBL_FONT
    c.fill = LBL_FILL
    c.alignment = CENTER
    c.border = BORDER

tributos = [
    ("COFINS",   COFINS_BASE, COFINS_ALIQ, COFINS_VAL),
    ("ICMS",     ICMS_BASE,   ICMS_ALIQ,   ICMS_VAL),
    ("PIS/PASEP", PIS_BASE,   PIS_ALIQ,    PIS_VAL),
]
for i, (n, b, a, v) in enumerate(tributos, start=41):
    ws2.cell(row=i, column=1, value=n).border = BORDER
    cb = ws2.cell(row=i, column=2, value=b); cb.number_format = BRL; cb.border = BORDER
    ca = ws2.cell(row=i, column=3, value=a); ca.number_format = '0.0000%'; ca.border = BORDER
    cv = ws2.cell(row=i, column=4, value=v); cv.number_format = BRL; cv.border = BORDER

# Bloco NFe
ws2["A45"] = "NOTA FISCAL ELETRÔNICA"
ws2["A45"].font = HDR_FONT
ws2["A45"].fill = HDR_FILL
ws2.merge_cells("A45:D45")
ws2["A45"].alignment = CENTER

pair(46, "Nº NFe",              NFE_NUMERO)
pair(47, "Série",               NFE_SERIE)
pair(48, "CFOP",                CFOP)
pair(49, "Chave de acesso",     CHAVE)
pair(50, "Protocolo autoriz.",  PROTOCOLO)

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 48
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14

# ============ Aba 3: Histórico de Consumo ============
ws3 = wb.create_sheet("Histórico")

ws3.merge_cells("A1:D1")
t = ws3["A1"]
t.value = "HISTÓRICO DE CONSUMO · últimos 13 meses"
t.font = TITLE_FONT
t.alignment = CENTER

hdr = ["Mês/Ano", "Consumo (kWh)", "Dias", "Tipo de faturamento"]
for i, h in enumerate(hdr, start=1):
    c = ws3.cell(row=2, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

for i, (m, k, d, t) in enumerate(HISTORICO, start=3):
    ws3.cell(row=i, column=1, value=m).border = BORDER
    cc = ws3.cell(row=i, column=2, value=k); cc.number_format = KWH; cc.border = BORDER
    ws3.cell(row=i, column=3, value=d).border = BORDER
    ws3.cell(row=i, column=4, value=t).border = BORDER
    for col in range(1, 5):
        ws3.cell(row=i, column=col).alignment = CENTER

# linha de média
mr = 3 + len(HISTORICO)
ws3.cell(row=mr, column=1, value="MÉDIA").font = LBL_FONT
ws3.cell(row=mr, column=1).fill = LBL_FILL
ws3.cell(row=mr, column=1).border = BORDER
cm = ws3.cell(row=mr, column=2, value=MEDIA)
cm.number_format = KWH
cm.font = LBL_FONT
cm.fill = LBL_FILL
cm.border = BORDER
for col in (3, 4):
    c = ws3.cell(row=mr, column=col, value="")
    c.fill = LBL_FILL
    c.border = BORDER

auto_width(ws3)
ws3.freeze_panes = "A3"

wb.save(OUT)
print(f"OK · gerado em {OUT}")
print(f"   - aba BASE: 1 registro no formato do etl_energia.py")
print(f"   - aba Detalhes: identificação, leitura, consumo, tributos, NFe")
print(f"   - aba Histórico: {len(HISTORICO)} meses + média")
