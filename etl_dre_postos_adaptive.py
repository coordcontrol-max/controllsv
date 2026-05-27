#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ETL DRE Postos (Adaptive) — parseia o export do relatório
'Mapa Anual de Resultados e Indicadores' do ERP Adaptive e gera
dados_dre_postos_adaptive/{ano}.json para o controllsv.web.app.

Fonte: o usuário exporta no Adaptive (Financeiro > Relatórios > Gerencial >
Mapa Anual de Resultados e Indicadores) TODOS os postos do ano corrente em
Excel. O arquivo cai em /mnt/c/Users/wesley/Downloads/. Esta máquina (WSL,
00CTR02) é o próprio PC e lê o Downloads direto.

Por que parsear o relatório e não ir no banco: o Adaptive é multi-tenant com
o Postgres em 127.0.0.1:6432 (inalcançável de fora) e a classificação 3.3.x
das despesas não está em campo nenhum do título (movimento_contabil e
lancamento_contabil estão vazios). O relatório já faz toda a classificação.
Ver memory/project_dre_postos_adaptive.md.

Uso:
  python3 etl_dre_postos_adaptive.py [caminho_do_xlsx]
Sem argumento, pega o mais recente '[postovivendas]Mapa Anual*.xlsx' do Downloads.
"""
import re, os, sys, glob, json, datetime as dt
import openpyxl

DOWNLOADS = "/mnt/c/Users/wesley/Downloads"
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados_dre_postos_adaptive")

MESES = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
         "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

# Linhas de topo da DRE (nível 0) — conjunto fixo
TOP = {
    "Total das Vendas", "(-) Custo Total", "(=) Lucro Bruto", "(-) Despesas",
    "(-) Devolução de Vendas", "(+) Receitas", "(=) Lucro Líquido",
    "(-) Retiradas", "(-) Investimentos", "Variação de Estoque", "(=) Saldo Final",
}
# Indicadores de topo (nível 0)
TOP_IND = {
    "Volume Vendido (Litros)", "Número Total de Abastecimentos",
    "Ticket Médio (R$)", "Ticket Médio (Litros)", "Faturamento por Litro",
    "Custo por Litro", "Lucro Bruto por Litro", "Despesa por Litro",
    "Lucro Líquido por Litro",
}

# código do posto (vw_empresa) -> trecho do nome que identifica
POSTO_CODIGOS = [
    ("001", "VS COMERCIAL"), ("002", "IRMAOS PACIFICOS"), ("003", "CIDADE OCIDENTAL"),
    ("004", "INFINITO"), ("005", "RIACHO FUNDO"), ("006", "SAMAMBAIA"),
    ("007", "EPTG"), ("008", "SM COMBUSTIVEIS"), ("009", "SETOR SUL"),
    ("010", "GM CENTRAL"), ("011", "SAO SEBASTIAO"),
]

CODE_RE = re.compile(r"^\d+\.\d+\.\d+")


def codigo_do_posto(nome):
    up = (nome or "").upper()
    for cod, frag in POSTO_CODIGOS:
        if frag in up:
            return cod
    return "999"


def is_noise(lab):
    if not lab:
        return True
    l = lab.strip()
    if l in ("", "Valores", "Subsessão", "Regime de Caixa"):
        return True
    if l.startswith("Empresa:") or l.startswith("[postovivendas]"):
        return True
    if l.startswith("Adaptive Business") or l.startswith("Página"):
        return True
    if l.startswith("Referência"):
        return True
    if set(l) <= {"-", " "}:
        return True
    return False


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def descobrir_ano(rows, fallback):
    for r in rows[:40]:
        for cell in r:
            if cell is None:
                continue
            m = re.search(r"Refer[êe]ncia[:\s]+(\d{4})", str(cell))
            if m:
                return int(m.group(1))
    return fallback


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    ano = descobrir_ano(rows, dt.date.today().year)
    postos = {}            # nome -> {"dre":[...], "indicadores":[...]}
    ordem = []             # ordem de aparição dos nomes
    cur_emp = None
    section = "dre"
    colmap = None
    cur_group = None
    cur_top = None

    for r in rows:
        lab = "" if not r or r[0] is None else str(r[0]).strip()

        if lab.startswith("Empresa:"):
            nome = lab.replace("Empresa:", "").strip()
            if nome != cur_emp:           # ignora repetição do header por página
                cur_emp = nome
                if nome not in postos:
                    postos[nome] = {"dre": [], "indicadores": []}
                    ordem.append(nome)
                section = "dre"; cur_group = None; cur_top = None
            continue

        if lab == "Subsessão":            # (re)mapeia meses pela posição
            cm = {}
            for ci, cell in enumerate(r):
                cv = "" if cell is None else str(cell).strip()
                if cv in MESES:
                    cm[MESES.index(cv)] = ci
                elif cv == "Total Anual":
                    cm["total"] = ci
            if cm:
                colmap = cm
            continue

        if lab == "Indicadores":
            section = "indicadores"; cur_group = None; cur_top = None
            continue

        if is_noise(lab) or cur_emp is None or colmap is None:
            continue

        valores = []
        for mi in range(12):
            ci = colmap.get(mi)
            valores.append(num(r[ci]) if ci is not None and ci < len(r) else None)
        total = num(r[colmap["total"]]) if "total" in colmap and colmap["total"] < len(r) else None

        topset = TOP if section == "dre" else TOP_IND
        if lab in topset:
            nivel = 0; cur_group = None; cur_top = lab
        elif section == "dre" and cur_top == "(-) Despesas":
            if CODE_RE.match(lab) or cur_group is None:
                nivel = 1; cur_group = lab
            else:
                nivel = 2
        else:
            nivel = 1; cur_group = None

        postos[cur_emp][section].append({
            "label": lab, "nivel": nivel,
            "grupo": cur_group if nivel == 2 else None,
            "meses": valores, "total": total,
        })

    lista_postos = [{"codigo": codigo_do_posto(n), "nome": n} for n in ordem]
    lista_postos.sort(key=lambda p: p["codigo"])
    return {
        "ano": ano,
        "geradoEm": dt.datetime.now().isoformat(timespec="seconds"),
        "fonte": "Mapa Anual de Resultados e Indicadores (Adaptive)",
        "arquivo": os.path.basename(path),
        "postos": lista_postos,
        "dados": postos,
    }


_DATA_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")


def parse_despesas_gerais(path):
    """Parseia o relatório 'Despesas Gerais' (Regime de Caixa) → títulos de
    despesa por posto → denominação (folha) → mês. É o detalhamento que casa
    centavo a centavo com as folhas do Mapa (mesmo regime). Colunas:
    0 Emissão · 1 Nome · 2 Nr Doc · 3 Denominação · 4 Observação · 5 Valor."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    ano = dt.date.today().year
    for r in rows[:30]:
        for cell in r:
            if cell and "Período" in str(cell):
                m = re.search(r"(\d{4})", str(cell))
                if m:
                    ano = int(m.group(1))
                break

    titulos = {}     # posto -> denominacao -> mes(str) -> [ {d,f,doc,o,v} ]
    cur_emp = None
    for r in rows:
        c0 = "" if not r or r[0] is None else str(r[0]).strip()
        if c0.startswith("Empresa:"):
            cur_emp = c0.replace("Empresa:", "").strip()
            titulos.setdefault(cur_emp, {})
            continue
        if cur_emp and _DATA_RE.match(c0):
            den = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
            val = num(r[5]) if len(r) > 5 else None
            if not den:
                continue
            mes = c0[3:5]
            rec = {
                "d": c0,
                "f": str(r[1]).strip() if len(r) > 1 and r[1] is not None else "",
                "doc": str(r[2]).strip() if len(r) > 2 and r[2] is not None else "",
                "o": str(r[4]).strip() if len(r) > 4 and r[4] is not None else "",
                "v": val,
            }
            titulos.setdefault(cur_emp, {}).setdefault(den, {}).setdefault(mes, []).append(rec)
    return {
        "ano": ano,
        "geradoEm": dt.datetime.now().isoformat(timespec="seconds"),
        "fonte": "Despesas Gerais — Regime de Caixa (Adaptive)",
        "arquivo": os.path.basename(path),
        "titulos": titulos,
    }


def gerar_despesas(ano_mapa):
    """Acha o último 'Despesas Gerais*.xlsx' no Downloads (prefere consolidado
    '[postovivendas]') e gera dados_dre_postos_adaptive/despesas_{ano}.json."""
    cands = glob.glob(os.path.join(DOWNLOADS, "*Despesas Gerais*.xlsx"))
    cands = [c for c in cands if not os.path.basename(c).startswith("~$")]
    if not cands:
        print("  · (sem 'Despesas Gerais*.xlsx' — drilldown ficará sem detalhe)")
        return
    bracket = [c for c in cands if os.path.basename(c).startswith("[")]
    pool = bracket if bracket else cands
    path = max(pool, key=os.path.getmtime)
    print("  fonte despesas:", path)
    data = parse_despesas_gerais(path)
    npost = len(data["titulos"])
    ndoc = sum(len(lst) for d in data["titulos"].values()
               for mm in d.values() for lst in mm.values())
    out = os.path.join(OUT_DIR, f"despesas_{data['ano']}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  ✓ gerado: {out} ({os.path.getsize(out)//1024} KB · {npost} posto(s) · {ndoc} títulos)")


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        cands = glob.glob(os.path.join(DOWNLOADS, "*Mapa Anual*.xlsx"))
        cands = [c for c in cands if not os.path.basename(c).startswith("~$")]
        if not cands:
            print("✗ Nenhum '*Mapa Anual*.xlsx' encontrado em", DOWNLOADS)
            sys.exit(1)
        # Prefere o export com TODOS os postos (prefixo "[postovivendas]"); só cai
        # pro export avulso (1 posto) se não houver o consolidado.
        bracket = [c for c in cands if os.path.basename(c).startswith("[")]
        pool = bracket if bracket else cands
        path = max(pool, key=os.path.getmtime)

    print("  fonte:", path)
    data = parse(path)
    print(f"  ano: {data['ano']} · postos: {len(data['postos'])}")

    os.makedirs(OUT_DIR, exist_ok=True)
    # Mapa Anual do Adaptive é em REGIME DE COMPETÊNCIA — vai pra arquivo próprio
    # ({ano}_competencia.json). O caixa fica em {ano}.json (gerado pelo SQL ETL).
    out = os.path.join(OUT_DIR, f"{data['ano']}_competencia.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print("  ✓ gerado:", out, f"({os.path.getsize(out)//1024} KB)")

    # Detalhamento (drilldown): a fonte oficial passou a ser o etl_dre_postos_sql.py
    # (3460 títulos · 11 postos, regime de caixa). Mapa Anual NÃO sobrescreve mais
    # despesas_*.json — o ETL caixa cuida disso depois ([4d] no run_etl.sh).
    # gerar_despesas(data["ano"])  # desativado: evita sobrescrita parcial


if __name__ == "__main__":
    main()
