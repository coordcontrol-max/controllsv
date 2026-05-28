#!/usr/bin/env python3
r"""ETL: Meta de VENDA por loja/mês (Operação) → Metas Manuais do projeto-comercial.

Fonte: \\10.61.1.13\controller\02 - SUPERMERCADOS\01 - FATURAMENTO\2026\<Mes>\
       "Aprovação de Metas*.xlsx", aba "APROVAÇÃO LOJA" / "Aprovação Loja v2".
       col 7 = código da loja (canônico, com poucos códigos antigos), col 11 = VENDA (META).

Grava em app_data.metas_manuais (Postgres do projeto-comercial, via RENDER_DATABASE_URL):
   metas_manuais["YYYY-MM"].operacao.lojas["<cod>"] = {"meta_venda": <float>}

Uso:
   python3 etl_metas_operacao_lojas.py            # dry-run (só imprime)
   python3 etl_metas_operacao_lojas.py --write    # grava no banco
"""
import os, sys, glob, json, datetime
import openpyxl

BASE = "/mnt/controller/02 - SUPERMERCADOS/01 - FATURAMENTO/2026"
MESES = {
    "Janeiro": "2026-01", "Fevereiro": "2026-02", "Março": "2026-03",
    "Abril": "2026-04", "Maio": "2026-05", "Junho": "2026-06",
}
# cod_loja canônicos (= venda_historico). Filtro de linhas válidas.
VALID = {5, 7, 10, 11, 13, 14, 16, 18, 20, 21, 23, 26, 27, 28, 29,
         101, 102, 103, 104, 106, 108, 109, 112, 117, 125, 131, 215, 219, 222}
# Códigos antigos que a planilha ainda usa em alguns meses → cod_loja canônico.
OLD2NEW = {9: 29, 17: 109}


def achar_arquivo(mes_nome):
    cand = [f for f in glob.glob(f"{BASE}/{mes_nome}/Aprova*Metas*.xls*")
            if "~$" not in os.path.basename(f)]
    return max(cand, key=os.path.getsize) if cand else None


def achar_aba(wb):
    for s in wb.sheetnames:
        if s.upper().strip() in ("APROVAÇÃO LOJA V2", "APROVAÇÃO LOJA"):
            return s
    for s in wb.sheetnames:
        u = s.upper()
        if "APROVA" in u and "LOJA" in u and "(2)" not in u:
            return s
    return None


def extrair_mes(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    aba = achar_aba(wb)
    if not aba:
        return None, {}
    metas = {}
    for row in wb[aba].iter_rows(min_row=5, max_row=60, values_only=True):
        cod = row[7] if len(row) > 7 else None
        mv = row[11] if len(row) > 11 else None
        if not isinstance(cod, (int, float)) or not isinstance(mv, (int, float)) or mv <= 0:
            continue
        c = OLD2NEW.get(int(cod), int(cod))
        if c not in VALID:
            continue
        if c not in metas:            # primeira ocorrência = linha-resumo da loja
            metas[c] = float(mv)
    return aba, metas


def coletar():
    out = {}
    for mes_nome, ym in MESES.items():
        path = achar_arquivo(mes_nome)
        if not path:
            print(f"  {ym} {mes_nome}: SEM ARQUIVO — pulando")
            continue
        aba, metas = extrair_mes(path)
        if not metas:
            print(f"  {ym} {mes_nome}: aba não encontrada / vazia — pulando")
            continue
        tot = sum(metas.values())
        falt = sorted(VALID - set(metas.keys()))
        print(f"  {ym} {mes_nome:<10} aba='{aba}' · {len(metas)} lojas · "
              f"meta venda TOTAL={tot:,.0f}" + (f" · sem meta: {falt}" if falt else ""))
        out[ym] = metas
    return out


def gravar(dados):
    import psycopg2
    env = {}
    here = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agente", ".env")
    for ln in open(here):
        ln = ln.strip()
        if "=" in ln and not ln.startswith("#"):
            k, v = ln.split("=", 1)
            env[k] = v
    con = psycopg2.connect(env["RENDER_DATABASE_URL"])
    cur = con.cursor()
    cur.execute("SELECT data FROM app_data WHERE key='metas_manuais'")
    r = cur.fetchone()
    allm = (r[0] if r and r[0] else {}) or {}
    hoje = datetime.date.today().isoformat()
    pulados = []
    for ym, metas in dados.items():
        mes = allm.setdefault(ym, {})
        # Respeita edição manual: se alguém ajustou a meta deste mês na tela
        # (Metas Manuais → Operação), não sobrescreve.
        if mes.get("operacao", {}).get("_fonte") == "manual":
            pulados.append(ym)
            continue
        mes["operacao"] = {
            "lojas": {str(c): {"meta_venda": v} for c, v in sorted(metas.items())},
            "_fonte": "aprovacao_metas",
            "_atualizado": hoje,
        }
    if pulados:
        print(f"  (pulados — meta manual: {', '.join(pulados)})")
    cur.execute(
        """INSERT INTO app_data (key, data, updated_by, updated_at)
           VALUES ('metas_manuais', %s::jsonb, NULL, NOW())
           ON CONFLICT (key) DO UPDATE SET data=EXCLUDED.data, updated_at=NOW()""",
        [json.dumps(allm, ensure_ascii=False)],
    )
    con.commit()
    cur.close()
    con.close()
    print(f"\n✓ gravado em app_data.metas_manuais · {len(dados)} meses · "
          f"operacao.lojas por mês")


def main():
    write = "--write" in sys.argv
    print(f"{'GRAVAÇÃO' if write else 'DRY-RUN'} — Meta de Venda Operação por loja/mês\n")
    dados = coletar()
    if not dados:
        print("\nNada coletado.")
        return
    if write:
        gravar(dados)
    else:
        print("\n(dry-run — rode com --write pra gravar no banco)")


if __name__ == "__main__":
    main()
