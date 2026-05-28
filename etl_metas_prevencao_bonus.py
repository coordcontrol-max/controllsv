#!/usr/bin/env python3
r"""ETL: histórico de metas do Bônus Prevenção → Metas Manuais (setor Prevenção).

Fonte: \\10.61.1.13\controller\02 - SUPERMERCADOS\00 - GESTÃO DE RESULTADOS\
       02 - PREVENÇÃO\2026\00 - Bônus Prevenção 2026.xlsx
       abas "Metas Janeiro".."Metas Junho" — cabeçalho "Pontos da Meta | Meta | ... | Peso".

Grava em app_data.metas_manuais["YYYY-MM"].prevencao:
   { indicadores: [{id,label,meta,peso,formato}], _fonte:'bonus_prevencao', _atualizado }
Edição manual na tela marca _fonte='manual' e o ETL para de sobrescrever o mês.

Uso:
   python3 etl_metas_prevencao_bonus.py            # dry-run
   python3 etl_metas_prevencao_bonus.py --write     # grava no banco
"""
import os, sys, json, datetime, unicodedata
import openpyxl

XLSX = ("/mnt/controller/02 - SUPERMERCADOS/00 - GESTÃO DE RESULTADOS/"
        "02 - PREVENÇÃO/2026/00 - Bônus Prevenção 2026.xlsx")
ABAS = {
    "Metas Janeiro": "2026-01", "Metas Fevereiro": "2026-02", "Metas Março": "2026-03",
    "Metas Abril": "2026-04", "Metas Maio": "2026-05", "Metas Junho": "2026-06",
}
PARAR = ("total", "obs:", "peso 1", "resultado")  # labels que encerram a tabela / cabeçalho


def _norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    return " ".join(s.split())


def id_canonico(label):
    n = _norm(label)
    if "uso/consumo" in n or "uso consumo" in n or "rotativo" in n:      return "inv_uso_consumo"
    if "contagem errada" in n:                                          return "inv_contagem_errada"
    if "inventario trocas" in n or "estoque das trocas" in n:           return "inv_trocas"
    if "trocas nao realizadas" in n:                                    return "trocas_nao_realizadas"
    if "troca sobre a venda" in n or "troca s/ venda" in n or n == "trocas": return "troca_sobre_venda"
    if "inventario desvio" in n or n == "desvio":                       return "inv_desvio"
    if "estoque negativo" in n:                                         return "estoque_negativo"
    if "cancelamento" in n:                                             return "cancelamento"
    if "furto evitado" in n or "evidencias de furto" in n:             return "furto_evitado"
    if "cftv" in n or "ocorrencias" in n:                               return "ocorrencias_cftv"
    if "resultado de perdas" in n:                                      return "resultado_perdas"
    if "inventario bovino" in n or "bovino" in n:                       return "inventario_bovino"
    if "quebra de inventario" in n:                                     return "quebra_inventario"
    return "ind_" + n.replace(" ", "_").replace("/", "_")[:24]


def _fmt(meta):
    if not isinstance(meta, (int, float)):
        return "qtd"
    if 0 < abs(meta) < 1:
        return "pct"
    if abs(meta) >= 100000:
        return "moeda"
    return "qtd"


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def achar_header(ws):
    """Retorna (row_idx, labelCol, metaCol, pesoCol) localizando 'Pontos da Meta'/'Meta'/'Peso'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        cols = {_norm(c): j for j, c in enumerate(row) if isinstance(c, str)}
        if "pontos da meta" in cols and "meta" in cols and "peso" in cols:
            return i, cols["pontos da meta"], cols["meta"], cols["peso"]
    return None


def extrair_aba(ws):
    h = achar_header(ws)
    if not h:
        return []
    hr, cL, cM, cP = h
    out = []
    vistos = set()
    for row in ws.iter_rows(min_row=hr + 1, max_row=hr + 25, values_only=True):
        label = row[cL] if len(row) > cL else None
        if not isinstance(label, str) or not label.strip():
            continue
        n = _norm(label)
        if n in PARAR or n.startswith("obs:") or n.startswith("peso 1"):
            break
        meta = _num(row[cM]) if len(row) > cM else None
        peso = _num(row[cP]) if len(row) > cP else None
        if meta is None and peso is None:
            continue
        iid = id_canonico(label)
        if iid in vistos:
            continue
        vistos.add(iid)
        out.append({"id": iid, "label": label.strip(), "meta": meta, "peso": peso,
                    "formato": _fmt(meta)})
    return out


def coletar():
    if not os.path.isfile(XLSX):
        print("SEM ARQUIVO:", XLSX)
        return {}
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    out = {}
    for aba, ym in ABAS.items():
        if aba not in wb.sheetnames:
            print(f"  {ym}: aba '{aba}' não existe — pulando")
            continue
        inds = extrair_aba(wb[aba])
        if not inds:
            print(f"  {ym}: nenhum indicador extraído — pulando")
            continue
        comMeta = sum(1 for i in inds if i["meta"] is not None)
        print(f"  {ym}  '{aba}': {len(inds)} indicadores ({comMeta} com meta)")
        out[ym] = inds
    return out


def gravar(dados):
    import psycopg2
    env = {}
    here = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agente", ".env")
    for ln in open(here):
        ln = ln.strip()
        if "=" in ln and not ln.startswith("#"):
            k, v = ln.split("=", 1)
            env[k] = v
    con = psycopg2.connect(env["RENDER_DATABASE_URL"])
    cur = con.cursor()
    cur.execute("SELECT data FROM app_data WHERE key='metas_manuais'")
    r = cur.fetchone()
    allm = (r[0] if r and r[0] else {}) or {}
    hoje = datetime.date.today().isoformat()
    pulados = []
    for ym, inds in dados.items():
        mes = allm.setdefault(ym, {})
        if mes.get("prevencao", {}).get("_fonte") == "manual":
            pulados.append(ym)
            continue
        mes["prevencao"] = {"indicadores": inds, "_fonte": "bonus_prevencao", "_atualizado": hoje}
    cur.execute(
        """INSERT INTO app_data (key, data, updated_by, updated_at)
           VALUES ('metas_manuais', %s::jsonb, NULL, NOW())
           ON CONFLICT (key) DO UPDATE SET data=EXCLUDED.data, updated_at=NOW()""",
        [json.dumps(allm, ensure_ascii=False)],
    )
    con.commit()
    cur.close()
    con.close()
    print(f"\n✓ gravado em app_data.metas_manuais · {len(dados)} meses · prevencao.indicadores")
    if pulados:
        print(f"  (pulados — meta manual: {', '.join(pulados)})")


def main():
    write = "--write" in sys.argv
    print(f"{'GRAVAÇÃO' if write else 'DRY-RUN'} — Metas Bônus Prevenção por mês\n")
    dados = coletar()
    if not dados:
        print("\nNada coletado.")
        return
    if write:
        gravar(dados)
    else:
        print("\n(dry-run — rode com --write pra gravar)")


if __name__ == "__main__":
    main()
