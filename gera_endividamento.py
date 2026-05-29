#!/usr/bin/env python3
r"""Endividamento & Contratos — monta o dataset do relatório de empréstimos.

Cruza:
  - CONTRATOS (autoritativo): PDFs em \\10.61.1.13\digitaliza\EMPRESTIMOS\CONTRATOS
    (montado em /mnt/digitaliza). Aqui usamos os termos consolidados da planilha
    de controle + linkamos os PDFs por empresa/banco; a planilha é CONFERIDA
    contra o pago real (divergências sinalizadas), não tratada como verdade.
  - PLANILHA: \\...\EMPRESTIMOS\EMPRESTIMOS ATUALIZADOS 2.0.xlsx
      aba RESUMO        = 1 linha por contrato (empresa, banco, tipo, data,
                          valor, juros% a.a., parcela, vencimentos, saldo devedor)
      abas por empresa  = cronograma de parcelas (nº, vencimento, valor,
                          juros proj/real, valor pago)
  - PAGO REAL (Oracle/DFC): rawOracle/{ano-mm}__fluxo_pago + __fluxo_juros.
      Empréstimo BANCÁRIO = CODESPECIE EMPRE2 (NOMERAZAO = banco). Agrega o
      principal (fluxo_pago) + juros (fluxo_juros) pago por banco/mês.

Grava em endividamento/atual no Firestore (projeto-686e2), consumido pela aba
"Endividamento & Contratos" da página Auditorias (Supervendas).

Uso (WSL, com /mnt/digitaliza montado):
  LD_LIBRARY_PATH=/opt/oracle/instantclient_23_5 python3 gera_endividamento.py [ano]
"""
import os
import sys
import glob
import re
import datetime as dt
from collections import defaultdict

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, "agente"))
import agente
import engine_fluxo
from firebase_admin import firestore

db = agente.db

XLSX = "/mnt/digitaliza/EMPRESTIMOS/EMPRESTIMOS ATUALIZADOS 2.0.xlsx"
PDF_DIR = "/mnt/digitaliza/EMPRESTIMOS/CONTRATOS"
# URL UNC pros links (abre no Windows do usuário)
PDF_UNC = r"\\10.61.1.13\digitaliza\EMPRESTIMOS\CONTRATOS"

# Empréstimo bancário no DFC = EMPRE2. Os demais (EMPRES/EMPREC/MUTPAG/MUTREC)
# são mútuos intercompany — fora do escopo de "contratos bancários".
COD_BANCARIO = {"EMPRE2"}
BANCOS_CANON = {
    "santander": "Santander", "banco brasil": "Banco do Brasil",
    "banco do brasil": "Banco do Brasil", "brasil": "Banco do Brasil",
    "itau": "Itaú", "itaú": "Itaú",
}


def _norm_banco(s):
    s = (str(s or "")).strip().lower()
    for k, v in BANCOS_CANON.items():
        if k in s:
            return v
    return (str(s).strip().title() or "—")


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def _iso(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
    return None


def _val(r, *ks):
    for k in ks:
        if k in r and r[k] is not None:
            return r[k]
        kl = k.lower()
        if kl in r and r[kl] is not None:
            return r[kl]
    return None


# ─── 1) Planilha: RESUMO (contratos) ─────────────────────────────────────────
def ler_resumo(wb):
    ws = wb["RESUMO"]
    contratos = []
    emp_atual = None
    seq = 0
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r:
            continue
        emp = (str(r[0]).strip() if r[0] else None)
        banco = r[1] if len(r) > 1 else None
        if emp:
            emp_atual = emp
        if not banco:          # linha vazia / separadora
            continue
        banco_norm = _norm_banco(banco)
        # só aceita linha de contrato real (banco conhecido); descarta rodapé/totais
        if banco_norm not in ("Santander", "Banco do Brasil", "Itaú"):
            continue
        if _num(r[4]) <= 0:    # sem valor → não é contrato
            continue
        # código da empresa = número líder ("01 - POLEN" → 1)
        m = re.match(r"\s*(\d+)", emp_atual or "")
        nro = int(m.group(1)) if m else None
        seq += 1
        contratos.append({
            "id": f"C{seq:02d}",
            "empresa": (emp_atual or "").strip(),
            "nroempresa": nro,
            "banco": _norm_banco(banco),
            "tipo": str(r[2] or "").strip(),
            "dataContratacao": _iso(r[3]),
            "valorTotal": _num(r[4]),
            "jurosAnualPct": round(_num(r[5]) * 100, 2),
            "valorJuros": _num(r[6]),
            "parcelaMensal": _num(r[8]) if len(r) > 8 else 0.0,
            "diaVenc": str(r[9] or "").strip() if len(r) > 9 else "",
            "totalParcelas": int(_num(r[10])) if len(r) > 10 else 0,
            "parcelasRemanescentes": int(_num(r[11])) if len(r) > 11 else 0,
            "data1aParcela": _iso(r[12]) if len(r) > 12 else None,
            "dataUltimaParcela": _iso(r[13]) if len(r) > 13 else None,
            "saldoDevedor": _num(r[15]) if len(r) > 15 else 0.0,
        })
    return contratos


# ─── 2) Planilha: abas por empresa (cronograma de parcelas) ──────────────────
def ler_cronogramas(wb):
    """tab '01-POLEN' → {nroempresa: [ {nro,venc,valor,jurosProj,jurosReal,pago}... ]}"""
    sched = {}
    for name in wb.sheetnames:
        m = re.match(r"\s*(\d+)\s*-", name)
        if not m:
            continue
        nro = int(m.group(1))
        ws = wb[name]
        parcelas = []
        cnpj = None
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not r:
                continue
            # metadados na coluna 9/10
            if len(r) > 9 and str(r[8] or "").strip().upper() == "CNPJ":
                cnpj = str(r[9] or "").strip()
            # parcela: col1 numérica
            if r[0] is None or not str(r[0]).strip().lstrip("-").isdigit():
                continue
            venc = _iso(r[1]) if len(r) > 1 else None
            parcelas.append({
                "nro": int(_num(r[0])),
                "venc": venc,
                "valor": _num(r[2]) if len(r) > 2 else 0.0,
                "jurosProj": _num(r[3]) if len(r) > 3 else 0.0,
                "jurosReal": _num(r[4]) if len(r) > 4 else 0.0,
                "pago": _num(r[5]) if len(r) > 5 else 0.0,
            })
        sched[nro] = {"cnpj": cnpj, "parcelas": parcelas}
    return sched


# ─── 3) PDFs dos contratos (links + presença) ────────────────────────────────
def listar_pdfs():
    """{nroempresa: [ {nome, tipo, banco, url} ]} a partir dos nomes dos arquivos."""
    out = defaultdict(list)
    if not os.path.isdir(PDF_DIR):
        print(f"   ⚠ {PDF_DIR} não acessível — pulando links de PDF")
        return out
    for path in sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf"))):
        nome = os.path.basename(path)
        m = re.match(r"\s*(\d+)", nome)
        nro = int(m.group(1)) if m else None
        up = nome.upper()
        tipo = ("CONTRATO" if "CONTRATO" in up else
                "CG" if " CG" in up or "-CG" in up else
                "PROJECAO" if "PROJE" in up else
                "CAPITALIZACAO" if "CAPITALIZ" in up else "OUTRO")
        out[nro].append({
            "nome": nome, "tipo": tipo,
            "banco": _norm_banco(nome),
            "url": PDF_UNC + "\\" + nome,
        })
    return out


# ─── 4) Pago real no Oracle (EMPRE2 principal + juros) por banco/mês ──────────
def _ler_slug(ano, mes, slug):
    """Lê rawOracle/{ano-mm}__{slug} (chunked) — só o slug pedido, sem carregar
    transitorias/opfin (1M+ linhas). Não toca no Oracle, só Firestore."""
    doc_id = f"{ano:04d}-{mes:02d}__{slug}"
    snap = db.collection("rawOracle").document(doc_id).get()
    if not snap.exists:
        return []
    data = snap.to_dict() or {}
    if not data.get("chunked"):
        return data.get("rows", [])
    rows = []
    chunks_ref = db.collection("rawOracle").document(doc_id).collection("chunks")
    for n in range(int(data.get("totalChunks") or 0)):
        cs = chunks_ref.document(str(n)).get()
        if cs.exists:
            rows.extend((cs.to_dict() or {}).get("rows", []))
    return rows


def pago_oracle_por_banco(ano):
    """Soma o pago bancário (EMPRE2) por banco e mês, do rawOracle já em Firestore.
    Retorna {banco: {"total": x, "porMes": {mm: v}, "juros": j}}."""
    res = defaultdict(lambda: {"total": 0.0, "juros": 0.0, "porMes": defaultdict(float)})
    for mes in range(1, 13):
        for fonte, campo in (("fluxo_pago", "total"), ("fluxo_juros", "juros")):
            for r in _ler_slug(ano, mes, fonte):
                if _val(r, "CODESPECIE") not in COD_BANCARIO:
                    continue
                banco = _norm_banco(_val(r, "NOMERAZAO"))
                v = abs(_num(_val(r, "VLROPERACAO")))
                res[banco][campo] += v
                if campo == "total":
                    res[banco]["porMes"][f"{mes:02d}"] += v
    # normaliza defaultdicts
    return {b: {"total": round(d["total"], 2), "juros": round(d["juros"], 2),
                "porMes": {k: round(v, 2) for k, v in d["porMes"].items()}}
            for b, d in res.items()}


def gerar(ano):
    print(f">> Endividamento {ano}…")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    contratos = ler_resumo(wb)
    sched = ler_cronogramas(wb)
    pdfs = listar_pdfs()
    print(f"   contratos (RESUMO): {len(contratos)} · cronogramas: {len(sched)} empresas · PDFs: {sum(len(v) for v in pdfs.values())}")

    nro2loja = {}
    ls = db.collection("meta").document("lojas").get().to_dict() or {}
    for it in (ls.get("items") or []):
        for n in (it.get("nroempresa") or []):
            try:
                nro2loja[int(n)] = it.get("descricao")
            except Exception:
                pass

    hoje = dt.date.today()
    # enriquece cada contrato com cronograma, pdfs, loja, pago da planilha
    for c in contratos:
        nro = c["nroempresa"]
        c["loja"] = nro2loja.get(nro)
        sc = sched.get(nro, {})
        parc = sc.get("parcelas", [])
        c["cnpj"] = sc.get("cnpj")
        c["pdfs"] = pdfs.get(nro, [])
        c["pagoPlanilha"] = round(sum(p["pago"] for p in parc), 2)
        c["pago2026"] = round(sum(p["pago"] for p in parc if p["venc"] and p["venc"][:4] == str(ano)), 2)
        c["jurosRealPlanilha"] = round(sum(p["jurosReal"] for p in parc), 2)
        # cronograma futuro (parcelas a vencer)
        futuras = [p for p in parc if p["venc"] and p["venc"] >= hoje.isoformat()]
        c["parcelasFuturas"] = futuras
        c["aVencer"] = round(sum(p["valor"] for p in futuras), 2)
        c["status"] = "Quitado" if c["saldoDevedor"] <= 0.01 else "Em aberto"

    pago_banco = pago_oracle_por_banco(ano)

    # KPIs
    ativos = [c for c in contratos if c["status"] != "Quitado"]
    total_contratado = round(sum(c["valorTotal"] for c in contratos), 2)
    saldo_total = round(sum(c["saldoDevedor"] for c in contratos), 2)
    juros_total = round(sum(c["valorJuros"] for c in contratos), 2)
    # custo médio ponderado pelo valor
    base = sum(c["valorTotal"] for c in ativos) or 1
    custo_medio = round(sum(c["jurosAnualPct"] * c["valorTotal"] for c in ativos) / base, 2)

    # composição por banco (saldo) e por tipo
    por_banco = defaultdict(float)
    por_tipo = defaultdict(float)
    for c in contratos:
        por_banco[c["banco"]] += c["saldoDevedor"]
        por_tipo[c["tipo"] or "—"] += c["saldoDevedor"]

    # cronograma de vencimentos por ano (parcelas futuras de todos)
    venc_ano = defaultdict(float)
    for c in contratos:
        for p in c["parcelasFuturas"]:
            venc_ano[p["venc"][:4]] += p["valor"]

    # reconciliação por banco: pago planilha × pago Oracle
    bancos = sorted(set(list(por_banco.keys()) + list(pago_banco.keys())))
    reconc = []
    for b in bancos:
        pl = round(sum(c["pago2026"] for c in contratos if c["banco"] == b), 2)
        ora = pago_banco.get(b, {}).get("total", 0.0)
        reconc.append({"banco": b, "pagoPlanilha": pl, "pagoOracle": ora,
                       "diferenca": round(pl - ora, 2)})

    doc = {
        "ano": ano,
        "geradoEm": firestore.SERVER_TIMESTAMP,
        "posicao": hoje.isoformat(),
        "kpis": {
            "totalContratado": total_contratado,
            "saldoDevedor": saldo_total,
            "jurosTotal": juros_total,
            "custoMedioAa": custo_medio,
            "contratosAtivos": len(ativos),
            "contratosTotal": len(contratos),
        },
        "contratos": contratos,
        "porBanco": [{"banco": k, "saldo": round(v, 2)} for k, v in sorted(por_banco.items(), key=lambda x: -x[1])],
        "porTipo": [{"tipo": k, "saldo": round(v, 2)} for k, v in sorted(por_tipo.items(), key=lambda x: -x[1])],
        "vencimentosPorAno": [{"ano": k, "valor": round(v, 2)} for k, v in sorted(venc_ano.items())],
        "pagoOraclePorBanco": pago_banco,
        "reconciliacao": reconc,
    }
    db.collection("endividamento").document("atual").set(doc, merge=False)
    print(f"✓ endividamento/atual: {len(contratos)} contratos, saldo R$ {saldo_total:,.2f}, "
          f"contratado R$ {total_contratado:,.2f}")
    print("   reconciliação por banco (planilha × Oracle):")
    for x in reconc:
        print(f"     {x['banco']:18} planilha {x['pagoPlanilha']:>14,.2f} · oracle {x['pagoOracle']:>14,.2f} · dif {x['diferenca']:>14,.2f}")


if __name__ == "__main__":
    gerar(int(sys.argv[1]) if len(sys.argv) > 1 else dt.date.today().year)
