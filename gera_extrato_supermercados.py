#!/usr/bin/env python3
"""Extrato Bancário SUPERMERCADO — por loja (AUTORIZADO DIRETORIA + SALDO PLANILHA).

Fonte: \\10.61.1.102\\exporta\\FINANCEIRO\\CONCILIAÇÃO\\PLANILHA DE CONCILIAÇÃO\\
       2026\\05 - MAIO\\DDMMAA - CONCILIAÇÃO BANCARIA.xlsx
(montado em /mnt/exporta). Um arquivo por dia útil.

Cada arquivo tem aba PAINEL com ~47 blocos (lojas + holdings) no formato:
  • [linha N]    "01-POLEN"
  • [linha N+1]  "AUTORIZADO DIRETORIA"  Banco | Conta | Valor | Destino
  • [linha N+2]  <valor autorizado>      SANTANDER | 13000793-6 | - | ...
  • [linha N+3..]  DEBITATO CC / TARIFAS / OUTROS PAG / PAG TESOURARIA / Dif
  • [linha N+10]  "SALDOS"   "R$ "  "DIFERENÇA"
  • [linha N+11]  "SALDO PLANILHA"   <valor saldo>
  • [linha N+12]  "SALDO CONCINCO"   ...

Saída: extratoBancario/{AAAA-MM} no Firestore (mesma coleção que postos/outras
usam pra pagamentoAutorizado/totalBancos GLOBAIS), mas adicionando porLoja:

  dias[DD] = {
    totalBancos:         <soma sobre lojas>,
    pagamentoAutorizado: <soma sobre lojas>,
    porLoja: {
      "L01": { totalBancos, pagamentoAutorizado },
      "L02": { ... },
      ...
    }
  }

Holdings (997-TIGO, 998-WITHI etc.) entram no porLoja com a tag literal
("L997"/"L998"); cabe ao consumidor (server.js) decidir se ignora.

Uso:
    python3 gera_extrato_supermercados.py [ANO] [MES]   # default = ano/mês corrente
"""
import os
import re
import sys
import warnings
import datetime as dt
import unicodedata
from collections import defaultdict
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

import firebase_admin
from firebase_admin import credentials, firestore

_HERE = os.path.dirname(os.path.abspath(__file__))
if not firebase_admin._apps:
    firebase_admin.initialize_app(credentials.Certificate(os.path.join(_HERE, "serviceAccount.json")))
db = firestore.client()

CONC_BASE = "/mnt/exporta/FINANCEIRO/CONCILIAÇÃO/PLANILHA DE CONCILIAÇÃO"
MESES_NOMES = ["01 - JANEIRO", "02 - FEVEREIRO", "03 - MARÇO", "04 - ABRIL",
               "05 - MAIO", "06 - JUNHO", "07 - JULHO", "08 - AGOSTO",
               "09 - SETEMBRO", "10 - OUTUBRO", "11 - NOVEMBRO", "12 - DEZEMBRO"]

# Regex pra extrair dia do nome do arquivo (DDMMAA + sufixo).
RE_DIA = re.compile(r"^(\d{2})(\d{2})(\d{2})\s*-\s*CONCILIA", re.I)
# Regex pra extrair o código L de loja do nome do bloco. Pega tanto "01-POLEN"
# (sem L explícito → vira "L01") quanto "997-TIGO" → "L997".
RE_LOJA = re.compile(r"^\s*L?\s*(\d{1,3})\s*[-–]\s*(.+?)\s*$", re.I)
LBL_AUTORIZ = "AUTORIZADO DIRETORIA"
LBL_SALDO   = "SALDO PLANILHA"


def _norm(s) -> str:
    if not s: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.upper().strip()


def _num(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace("R$", "").replace(".", "").replace(",", ".").strip()
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _carrega_mapa_nroemp_loja() -> dict[str, str]:
    """nroempresa (string) → loja descricao (ex: "L01"). Lê meta/lojas."""
    items = (db.collection("meta").document("lojas").get().to_dict() or {}).get("items", [])
    out = {}
    for x in items:
        desc = (x.get("descricao") or "").strip()
        for nro in (x.get("nroempresa") or []):
            out[str(nro).strip()] = desc
    return out


# Textos no Destino que devem ser IGNORADOS pro comentário (são corriqueiros,
# não são observações de divergência). Tudo o que NÃO casar com esses padrões
# E não estiver vazio vira observação a subir no comentário.
_DEST_SKIP_PATTERNS = (
    "TARIFA", "TRANSFER", "NAO HOUVE", "NÃO HOUVE",
    "PROTEGE REALIZADA", "PROTEGE CASH",
)

# Labels da coluna esquerda do bloco (categoria de pagamento) que viram linhas
# de "categoria" no comentário, com o valor da linha SEGUINTE em col c.
# Match exato após normalização (NFKD + upper + strip).
# Tupla: (chave_match, nome_exibido, signo).
# Convenção do usuário:
#   - TARIFAS / OUTROS PAG: positivo no Excel = debitado → INVERTER (signo -1).
#   - PAG TESOURARIA: negativo no Excel = debitado → INVERTER (signo -1) pra
#     sempre sair POSITIVO no comentário (ajuste pedido pelo usuário).
#   - TRANSFER PROTEGE: convenção não especificada → manter (raro como label).
_CAT_LABELS = (
    ("TARIFAS",                 "Tarifas",                -1),
    ("OUTROS PAG",              "Outros Pag.",            -1),
    ("PAG TESOURARIA",          "Pag. Tesouraria",        -1),
    ("TRANSFER PROTEGE",        "Transferência Protege",  +1),
    ("TRANSFERENCIA PROTEGE",   "Transferência Protege",  +1),
    ("TRANSF PROTEGE",          "Transferência Protege",  +1),
    ("TRANSFER. PROTEGE",       "Transferência Protege",  +1),
)


def _is_obs_relevante(destino: str) -> bool:
    if not destino: return False
    n = _norm(destino).strip()
    if not n or n == "-": return False
    # Exceção: qualquer linha com "PROTEGE" no destino passa SEMPRE — caso
    # típico é "TRANSFERENCIA PROTEGE REALIZADA" com valor ≠ 0 que cairia no
    # skip "TRANSFER" / "PROTEGE REALIZADA" se não tratada à parte.
    if "PROTEGE" in n:
        return True
    for p in _DEST_SKIP_PATTERNS:
        if p in n: return False
    return True


def _extrai_blocos(ws, mapa_nroemp_loja: dict[str, str]) -> list[dict]:
    """Lê todos os blocos da aba PAINEL → [{loja, autorizado, saldo, observacoes}].
    Além de autorizado/saldo, agrega observações relevantes do col Destino
    (pula TARIFA, TRANSFERENCIA PROTEGE, NÃO HOUVE TARIFA — só observações reais
    como REAL VLR. DIVERGENTE, ESTORNO EBEG, etc.) pra subir como comentário
    automático na linha "Diferenças Pag Autorizado" do DFC Consolidado."""
    blocos = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v: continue
            if "AUTORIZADO DIRETOR" not in _norm(v):
                continue
            nome = ws.cell(r - 1, c).value
            autoriz = _num(ws.cell(r + 1, c).value)
            saldo = 0.0
            # Acha o limite inferior do bloco — onde "SALDO PLANILHA" aparece.
            saldo_row = None
            for rr in range(r + 5, min(r + 20, ws.max_row + 1)):
                lbl = ws.cell(rr, c).value
                if lbl and "SALDO PLANILHA" in _norm(lbl):
                    saldo = _num(ws.cell(rr, c + 2).value)
                    saldo_row = rr
                    break
            mat = RE_LOJA.match(str(nome or ""))
            if not mat: continue
            nroemp = mat.group(1).lstrip("0") or "0"
            loja_cod = mapa_nroemp_loja.get(nroemp, "OUTROS")
            # Coleta as observações do col Destino (c+5 = col "Destino" no layout)
            # + o valor da mesma linha (c+4 = col "Valor"), nas linhas entre o
            # header (r) e o saldo (saldo_row). Também varre a coluna esquerda
            # (c) procurando labels de categoria (TARIFAS/OUTROS PAG/PAG
            # TESOURARIA/TRANSFER PROTEGE) — o valor vem da linha seguinte.
            obs = []
            categorias = {}    # nome amigável → valor (somando se aparecer >1x)
            r_fim = saldo_row if saldo_row else r + 10
            valor_col   = c + 4      # Banco c+2, Conta c+3, Valor c+4, Destino c+5
            destino_col = c + 5
            seen_txt = set()
            for rr in range(r + 1, r_fim):
                # 1) Destino vermelho (linhas de observação) na coluna direita.
                # Convenção (definida pelo usuário): MANTER o sinal nativo do
                # Excel pra qualquer linha da coluna Destino. As lojas usam
                # convenções diferentes pra "VLR DIVERGENTE" e similares, então
                # inverter à força gerava resultados conceitualmente errados
                # (ex.: REAL VLR DIVERGENTE L14 dia 18 vinha +2.953 quando o
                # próprio Excel já tinha -2.953).
                destino = ws.cell(rr, destino_col).value
                if _is_obs_relevante(destino):
                    txt = str(destino).strip()
                    val = _num(ws.cell(rr, valor_col).value)
                    if txt not in seen_txt and val:
                        seen_txt.add(txt)
                        obs.append({"texto": txt, "valor": val})
                # 2) Label de categoria na coluna esquerda. Sinal varia por
                # categoria (ver _CAT_LABELS).
                lbl = ws.cell(rr, c).value
                if lbl:
                    nlbl = _norm(str(lbl)).strip()
                    for chave, nome, signo in _CAT_LABELS:
                        if chave == nlbl and rr + 1 <= r_fim:
                            v = _num(ws.cell(rr + 1, c).value) * signo
                            if v:
                                categorias[nome] = categorias.get(nome, 0.0) + v
                            break
            blocos.append({"loja": loja_cod, "nroempresa": nroemp,
                            "nome_raw": str(nome).strip(),
                            "autorizado": autoriz, "saldo": saldo,
                            "observacoes": obs, "categorias": categorias})
    return blocos


def _processa_dia(path: str, mapa_nroemp_loja: dict[str, str]) -> tuple[str | None, dict, dict, dict]:
    """Retorna ('DD', {loja: {totalBancos, pagamentoAutorizado}}, {loja: [obs...]},
    {loja: {categoria_nome: valor}}). Agrupa por descricao da loja."""
    fname = os.path.basename(path)
    mat = RE_DIA.match(fname)
    if not mat: return None, {}, {}, {}
    dd = mat.group(1)
    wb = load_workbook(path, data_only=True)
    if "PAINEL" not in wb.sheetnames: return dd, {}, {}, {}
    blocos = _extrai_blocos(wb["PAINEL"], mapa_nroemp_loja)
    porLoja = {}
    obsPorLoja = {}
    catPorLoja = {}
    for b in blocos:
        l = b["loja"]
        if l in porLoja:
            porLoja[l]["totalBancos"]         += b["saldo"]
            porLoja[l]["pagamentoAutorizado"] += b["autorizado"]
        else:
            porLoja[l] = {"totalBancos": b["saldo"], "pagamentoAutorizado": b["autorizado"]}
        for o in b.get("observacoes", []):
            obsPorLoja.setdefault(l, []).append(o)
        for nome, val in (b.get("categorias") or {}).items():
            catPorLoja.setdefault(l, {})
            catPorLoja[l][nome] = catPorLoja[l].get(nome, 0.0) + val
    return dd, porLoja, obsPorLoja, catPorLoja


def gerar(ano: int, mes: int) -> None:
    pasta = os.path.join(CONC_BASE, str(ano), MESES_NOMES[mes - 1])
    if not os.path.isdir(pasta):
        print(f"✗ pasta não existe: {pasta}")
        return
    arquivos = sorted([f for f in os.listdir(pasta)
                       if f.lower().endswith(".xlsx") and not f.startswith("~$")])
    if not arquivos:
        print(f"  (sem arquivos em {pasta})"); return

    # Lê o doc atual (pra preservar dias gerados por outros ETLs).
    chave = f"{ano:04d}-{mes:02d}"
    ref = db.collection("extratoBancario").document(chave)
    doc_atual = ref.get().to_dict() or {}
    dias_atual = doc_atual.get("dias", {}) or {}

    print(f"\n>>> Extrato Supermercado {chave} — {len(arquivos)} arquivo(s)")
    mapa = _carrega_mapa_nroemp_loja()
    print(f"  mapa nroempresa→loja: {len(mapa)} entradas (ex: 1→{mapa.get('1')}, 101→{mapa.get('101')})")

    # Comentários auto na linha difPagAutorizado do DFC Consolidado.
    # Coleção comentariosCelula/{AAAA-MM}.cells[difPagAutorizado|DD|loja].
    ref_cmt = db.collection("comentariosCelula").document(chave)
    cmt_doc = ref_cmt.get().to_dict() or {}
    cells = (cmt_doc.get("cells") or {})
    cells_updates = {}   # patches a aplicar
    OBS_COUNT_TOTAL = 0

    def _fmt(v: float) -> str:
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _texto_com_total(linhas: list[str], total: float, limit: int = 5000) -> str:
        """Junta as linhas + 'TOTAL: …' GARANTINDO que o total fique no final.
        Se o body estourar `limit`, trunca o body (mantendo o TOTAL intacto)."""
        total_line = f"\nTOTAL: {_fmt(total)}"
        body = "\n".join(linhas)
        budget = limit - len(total_line) - 4   # margem pra "\n…"
        if len(body) > budget:
            body = body[:budget].rstrip() + "\n…"
        return body + total_line

    for fname in arquivos:
        path = os.path.join(pasta, fname)
        try:
            dd, porLoja, obsPorLoja, catPorLoja = _processa_dia(path, mapa)
        except Exception as e:
            print(f"  ✗ {fname}: {e}"); continue
        if not dd or not porLoja:
            print(f"  ⚠ {fname}: sem dado"); continue
        total_bancos = sum(v["totalBancos"] for v in porLoja.values())
        autoriz_total = sum(v["pagamentoAutorizado"] for v in porLoja.values())
        antigo = dias_atual.get(dd, {}) or {}
        dias_atual[dd] = {
            **antigo,
            "totalBancos": round(total_bancos, 2),
            "pagamentoAutorizado": round(autoriz_total, 2),
            "porLoja": {k: {"totalBancos": round(v["totalBancos"], 2),
                              "pagamentoAutorizado": round(v["pagamentoAutorizado"], 2)}
                          for k, v in porLoja.items()},
        }
        # Lojas que têm algo a comentar (categorias ou obs vermelhas).
        lojas_com_cmt = set(obsPorLoja.keys()) | set(catPorLoja.keys())

        # Comentários por loja (não sobrescreve edições manuais — só atualiza
        # entries marcadas com _auto=True ou inexistentes).
        for loja in lojas_com_cmt:
            obs_list = obsPorLoja.get(loja, [])
            cat_dict = catPorLoja.get(loja, {})
            if not obs_list and not cat_dict: continue
            cell_key = f"difPagAutorizado|{dd}|{loja}"
            cur = cells.get(cell_key) or {}
            if cur and not cur.get("_auto"):
                continue   # comentário manual — não toca
            linhas = []
            total = 0.0
            # Categorias primeiro (TARIFAS, OUTROS PAG, PAG TESOURARIA, TRANSFER PROTEGE)
            for nome in ("Tarifas", "Outros Pag.", "Pag. Tesouraria", "Transferência Protege"):
                if nome in cat_dict:
                    v = cat_dict[nome]
                    linhas.append(f"{nome} — {_fmt(v)}")
                    total += v
            # Depois as observações vermelhas.
            for o in obs_list:
                linhas.append(f"{o['texto']} — {_fmt(o['valor'])}")
                total += o["valor"]
            texto = _texto_com_total(linhas, total)
            cells_updates[cell_key] = {
                "texto": texto, "autor": "auto (conciliação)",
                "em": dt.datetime.now().isoformat(timespec="seconds"),
                "_auto": True,
            }
            OBS_COUNT_TOTAL += len(obs_list) + len(cat_dict)
        # Comentário consolidado/global (loja vazia): junta tudo do dia
        # ordenando por loja, pra quem olha "Todas as lojas".
        if lojas_com_cmt:
            todas = []
            total_g = 0.0
            for loja in sorted(lojas_com_cmt):
                cat_dict = catPorLoja.get(loja, {})
                for nome in ("Tarifas", "Outros Pag.", "Pag. Tesouraria", "Transferência Protege"):
                    if nome in cat_dict:
                        v = cat_dict[nome]
                        todas.append(f"[{loja}] {nome} — {_fmt(v)}")
                        total_g += v
                for o in obsPorLoja.get(loja, []):
                    todas.append(f"[{loja}] {o['texto']} — {_fmt(o['valor'])}")
                    total_g += o["valor"]
            if todas:
                k_glob = f"difPagAutorizado|{dd}|"
                cur_g = cells.get(k_glob) or {}
                if not cur_g or cur_g.get("_auto"):
                    cells_updates[k_glob] = {
                        "texto": _texto_com_total(todas, total_g),
                        "autor": "auto (conciliação)",
                        "em": dt.datetime.now().isoformat(timespec="seconds"),
                        "_auto": True,
                    }
        print(f"  ✓ dia {dd}: {len(porLoja)} lojas, totalBancos R$ {total_bancos:,.2f}, autorizDir R$ {autoriz_total:,.2f}, obs={sum(len(o) for o in obsPorLoja.values())}")

    if cells_updates:
        ref_cmt.set({"cells": cells_updates, "atualizadoEm": firestore.SERVER_TIMESTAMP}, merge=True)
        print(f"\n✓ comentariosCelula/{chave}: {len(cells_updates)} células atualizadas ({OBS_COUNT_TOTAL} obs)")

    ref.set({
        "ano": ano, "mes": mes,
        "dias": dias_atual,
        "atualizadoEm": firestore.SERVER_TIMESTAMP,
        "origem": "gera_extrato_supermercados.py (per-loja)",
    }, merge=True)
    print(f"\n✓ extratoBancario/{chave}: {len(dias_atual)} dias gravados")


if __name__ == "__main__":
    hoje = dt.date.today()
    ano = int(sys.argv[1]) if len(sys.argv) > 1 else hoje.year
    mes = int(sys.argv[2]) if len(sys.argv) > 2 else hoje.month
    gerar(ano, mes)
