#!/usr/bin/env python3
"""Pagamento Autorizado + Total Bancos POSTOS/OUTRAS — espelha gera_extrato_supermercados.

Fonte: \\10.61.1.13\\cvl\\CONTAS A PAGAR\\PLANILHA DE CONCILIAÇÃO\\AAAA
       (montada em /mnt/cvl). Estrutura: AAAA/MM - MES/DD - DIA/<arquivos.xlsx>.
       Em cada pasta de dia há um par:
         • CONCILIAÇÃO BANCARIA - POSTOS.xlsx
         • CONCILIAÇÃO BANCARIA -TTR.xlsx
       Em ambos consideramos SÓ a aba PAINEL (a Planilha2/LOJAS = supermercados,
       que tem ETL próprio em gera_extrato_supermercados.py).

Cada PAINEL é uma sequência de BLOCOS por entidade, no mesmo layout dos supermercados:

  • [linha N]    "AUTO POSTO XYZ"  (nome da entidade)
  • [linha N+1]  "AUTORIZADO DIRETORIA"  Banco | Conta | Valor | Destino
  • [linha N+2]  <valor autorizado>      SANTANDER | 1234-5 | - | ...
  • [linha N+3..]  TARIFAS / OUTROS PAG / PAG TESOURARIA / TRANSFER PROTEGE / DESTINO obs
  • [linha ~+10]  "SALDOS"   "R$ "  "DIFERENÇA"
  • [linha ~+11]  "SALDO PLANILHA"   <valor saldo>
  • [linha ~+12]  "SALDO CONCINCO"   ...

Classificação (espelhando a versão anterior):
  • Entidades em OUTRAS_NAMES (FLUXO/LP/PEGUI/RETA/TARES/TTR) → segmento "outras"
  • Demais (AUTO POSTO *, SETOR SUL, GM CENTRAL, …)            → segmento "postos"

Saída por dia/segmento:
  dados_fluxo_{seg}/AAAA-MM.json
    bancos[DD] = {
      pagamentoAutorizado:      <soma do segmento>,         (antigo)
      totalBancos:              <soma SALDO PLANILHA>,      (NOVO — espelha sup.)
      diferencasPagAutorizado:  <pagtoDia(dia) + autoriz>,  (antigo)
      porEntidade: {
        "AUTO POSTO XYZ": {
          pagamentoAutorizado, totalBancos, observacoes:[...], categorias:{...}
        },
        ...
      }
    }

Comentários automáticos:
  comentariosDFC/{AAAA-MM}-{seg}.difPagAutorizado.DD =
    "Tarifas — R$ X / Outros Pag. — R$ Y / [ENTIDADE] REAL VLR DIVERGENTE — R$ Z / TOTAL: R$ T"
  (mesmo formato do supermercados; chave segmentada pra não colidir com supermercados).
  Não sobrescreve comentários manuais (entries sem `_auto:true` ficam intactas).
"""
from __future__ import annotations
import os
import re
import sys
import json
import warnings
import datetime as dt
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

import firebase_admin
from firebase_admin import credentials, firestore

_HERE = os.path.dirname(os.path.abspath(__file__))
if not firebase_admin._apps:
    firebase_admin.initialize_app(credentials.Certificate(os.path.join(_HERE, "serviceAccount.json")))
db = firestore.client()

CONC_MOUNT = "/mnt/cvl/CONTAS A PAGAR/PLANILHA DE CONCILIAÇÃO"
CONC_LOCAL = "/tmp/concil_postos"
PROJ       = "/root/projeto_dre"
OUT_DIRS   = {"postos": os.path.join(PROJ, "dados_fluxo_postos"),
              "outras": os.path.join(PROJ, "dados_fluxo_outras")}

FILES = ["CONCILIAÇÃO BANCARIA - POSTOS.xlsx", "CONCILIAÇÃO BANCARIA -TTR.xlsx"]

# Agrupamentos que compõem o (−) Pagamentos (dia) de cada DFC Consolidado.
PAGTO_AGRUPS = {
    "postos": {"Fornecedores", "Pagto Entre Unidades", "Despesas"},
    "outras": {"Despesas"},
}

OUTRAS_NAMES = {"RETA", "COMERCIAL TARES", "TARES", "FLUXO", "LP", "PEGUI", "TTR"}

LBL_AUTORIZ = "AUTORIZADO DIRETORIA"
LBL_SALDO   = "SALDO PLANILHA"
LBL_TOTAL   = "TOTAL DE OPERACOES"     # ignorado (rodapé)

# Destino patterns ignorados pro comentário (espelha supermercados).
_DEST_SKIP_PATTERNS = (
    "TARIFA", "TRANSFER", "NAO HOUVE", "NÃO HOUVE",
    "PROTEGE REALIZADA", "PROTEGE CASH",
)

# Labels-categoria na coluna esquerda do bloco. Tupla: (chave_match, nome_exibido, signo).
# Convenções idênticas ao supermercados (definidas pelo user em iterações anteriores):
#   TARIFAS / TRANSFER PROTEGE: sinal nativo (+1)
#   OUTROS PAG / PAG TESOURARIA: INVERTER (-1) — sai sempre POSITIVO no comentário
_CAT_LABELS = (
    ("TARIFAS",                 "Tarifas",                +1),
    ("OUTROS PAG",              "Outros Pag.",            -1),
    ("PAG TESOURARIA",          "Pag. Tesouraria",        -1),
    ("TRANSFER PROTEGE",        "Transferência Protege",  +1),
    ("TRANSFERENCIA PROTEGE",   "Transferência Protege",  +1),
    ("TRANSF PROTEGE",          "Transferência Protege",  +1),
    ("TRANSFER. PROTEGE",       "Transferência Protege",  +1),
)


def _norm(s) -> str:
    if not s: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = re.sub(r"\s+", " ", s).strip().upper()
    s = re.sub(r"\s+\d{1,2}/\d{1,2}/\d{2,4}$", "", s)   # tira data no fim do header
    return s


def _num(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace("R$", "").replace(".", "").replace(",", ".").strip()
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _fmt(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _texto_com_total(linhas: list[str], total: float, limit: int = 50000) -> str:
    """Junta linhas + 'TOTAL:' no final. Trunca o body se estourar (mantém o total)."""
    total_line = f"\nTOTAL: {_fmt(total)}"
    body = "\n".join(linhas)
    budget = limit - len(total_line) - 4
    if len(body) > budget:
        body = body[:budget].rstrip() + "\n…"
    return body + total_line


def classifica(ent_norm: str) -> str:
    for nome in OUTRAS_NAMES:
        if ent_norm == nome or ent_norm.startswith(nome + " "):
            return "outras"
    return "postos"


def _is_obs_relevante(destino: str) -> bool:
    if not destino: return False
    n = _norm(destino)
    if not n or n == "-": return False
    if "PROTEGE" in n: return True   # ex.: TRANSFERENCIA PROTEGE REALIZADA com valor
    for p in _DEST_SKIP_PATTERNS:
        if p in n: return False
    return True


def _extrai_blocos(ws) -> list[dict]:
    """Lê todos os blocos da aba PAINEL.
    Retorna: [{nome_raw, ent_norm, autorizado, saldo, observacoes:[{texto,valor}], categorias:{nome:valor}}]."""
    blocos = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v: continue
            if _norm(v) != _norm(LBL_AUTORIZ):
                continue
            nome = ws.cell(r - 1, c).value if r > 1 else None
            ent_norm = _norm(nome)
            if not ent_norm or ent_norm == LBL_TOTAL:
                continue

            autoriz = _num(ws.cell(r + 1, c).value)

            # acha SALDO PLANILHA no bloco
            saldo = 0.0
            saldo_row = None
            for rr in range(r + 5, min(r + 20, ws.max_row + 1)):
                lbl = ws.cell(rr, c).value
                if lbl and LBL_SALDO in _norm(lbl):
                    saldo = _num(ws.cell(rr, c + 2).value)
                    saldo_row = rr
                    break

            # observações da col Destino (c+5) + categorias da col esquerda (c)
            obs = []
            categorias = {}
            r_fim = saldo_row if saldo_row else r + 10
            valor_col   = c + 4
            destino_col = c + 5
            seen_txt = set()
            for rr in range(r + 1, r_fim):
                destino = ws.cell(rr, destino_col).value
                if _is_obs_relevante(destino):
                    txt = str(destino).strip()
                    val = _num(ws.cell(rr, valor_col).value)
                    if txt not in seen_txt and val:
                        seen_txt.add(txt)
                        obs.append({"texto": txt, "valor": val})
                lbl = ws.cell(rr, c).value
                if lbl:
                    nlbl = _norm(lbl)
                    for chave, nome_exib, signo in _CAT_LABELS:
                        if chave == nlbl and rr + 1 <= r_fim:
                            v2 = _num(ws.cell(rr + 1, c).value) * signo
                            if v2:
                                categorias[nome_exib] = categorias.get(nome_exib, 0.0) + v2
                            break

            blocos.append({
                "nome_raw":   str(nome).strip(),
                "ent_norm":   ent_norm,
                "autorizado": autoriz,
                "saldo":      saldo,
                "observacoes": obs,
                "categorias":  categorias,
            })
    return blocos


def mes_de_pasta(nome: str):
    m = re.match(r"\s*(\d{1,2})", nome)
    return int(m.group(1)) if m else None


def dia_de_pasta(nome: str):
    m = re.match(r"\s*(\d{1,2})", nome)
    return int(m.group(1)) if m else None


def stage_local(refresh=False):
    """cópia local do mês via cp -r (drvfs 9p é flaky com muitos open())."""
    import subprocess
    if refresh or not os.path.isdir(CONC_LOCAL):
        if not os.path.isdir(CONC_MOUNT):
            return None
        os.makedirs(CONC_LOCAL, exist_ok=True)
        print(f"Copiando conciliação p/ {CONC_LOCAL} …")
        subprocess.run(["cp", "-r", f"{CONC_MOUNT}/.", CONC_LOCAL + "/"], check=True)
    return CONC_LOCAL


def coleta(base, ano_filtro=None, mes_filtro=None):
    """Varre a árvore. Retorna por_dia[(ano,mes,seg)][dia] = lista_de_blocos."""
    por_dia = defaultdict(lambda: defaultdict(list))   # (ano,mes,seg) -> {dia: [bloco]}

    # base = .../CONCILIAÇÃO/ ; estrutura esperada: base/AAAA/MM - MES/DD - DIA/*.xlsx
    # Mas também aceitamos base = .../AAAA quando stage_local copiou só um ano.
    raiz_anos = []
    for x in sorted(os.listdir(base)):
        full = os.path.join(base, x)
        if os.path.isdir(full) and re.match(r"^\d{4}$", x):
            raiz_anos.append((int(x), full))
    if not raiz_anos:
        # base já é um ano (caso do CONC_LOCAL=/tmp/concil_postos copiado do dir AAAA)
        raiz_anos = [(ano_filtro or dt.date.today().year, base)]

    for ano, ano_dir in raiz_anos:
        if ano_filtro and ano != ano_filtro: continue
        for mes_pasta in sorted(os.listdir(ano_dir)):
            mes_dir = os.path.join(ano_dir, mes_pasta)
            if not os.path.isdir(mes_dir): continue
            mes = mes_de_pasta(mes_pasta)
            if not mes: continue
            if mes_filtro and mes != mes_filtro: continue
            for dia_pasta in sorted(os.listdir(mes_dir)):
                dia_dir = os.path.join(mes_dir, dia_pasta)
                if not os.path.isdir(dia_dir): continue
                dia = dia_de_pasta(dia_pasta)
                if not dia: continue
                for fname in FILES:
                    fpath = os.path.join(dia_dir, fname)
                    if not os.path.exists(fpath) or os.path.basename(fpath).startswith("~$"):
                        continue
                    try:
                        wb = load_workbook(fpath, data_only=True)
                        if "PAINEL" not in wb.sheetnames:
                            wb.close(); continue
                        blocos = _extrai_blocos(wb["PAINEL"])
                        wb.close()
                    except Exception as e:
                        print(f"  ! erro lendo {fpath}: {e}")
                        continue
                    for b in blocos:
                        seg = classifica(b["ent_norm"])
                        por_dia[(ano, mes, seg)][dia].append(b)
    return por_dia


def pagto_dia_map(doc, seg):
    """{dia_int: pagtoDia} a partir de porAgrupamento (= (−) Pagamentos do dia)."""
    dim = doc.get("dim", {})
    agr_names = dim.get("agrupamentos", [])
    alvo_idx = {i for i, n in enumerate(agr_names) if n in PAGTO_AGRUPS[seg]}
    out = defaultdict(float)
    for e in doc.get("porAgrupamento", []):
        if e.get("a") in alvo_idx:
            out[e["d"]] += e.get("v", 0.0)
    return out


def aplica(por_dia, dry_run=False):
    """Grava JSON local + comentariosDFC segmentado no Firestore."""
    relatorio = []
    # cache de docs comentariosDFC por (seg, ano, mes) — chave doc = "AAAA-MM-{seg}"
    cmt_cache = {}
    cmt_updates = defaultdict(dict)   # docId → {linha.dia: novo_texto}

    for (ano, mes, seg) in sorted(por_dia):
        path = os.path.join(OUT_DIRS[seg], f"{ano:04d}-{mes:02d}.json")
        if not os.path.exists(path):
            print(f"  · sem JSON p/ {seg} {ano}-{mes:02d} (pulado: {path})")
            continue
        with open(path, encoding="utf-8") as f:
            doc = json.load(f)
        bancos = doc.get("bancos") or {}
        pagtos = pagto_dia_map(doc, seg)

        cmt_chave = f"{ano:04d}-{mes:02d}-{seg}"
        if cmt_chave not in cmt_cache:
            ref = db.collection("comentariosDFC").document(cmt_chave)
            cmt_cache[cmt_chave] = ref.get().to_dict() or {}
        cmt_doc = cmt_cache[cmt_chave]
        cmt_diff = (cmt_doc.get("difPagAutorizado") or {})

        for dia, blocos in sorted(por_dia[(ano, mes, seg)].items()):
            dd = f"{dia:02d}"
            # totais consolidados
            autoriz_total = sum(b["autorizado"] for b in blocos)
            saldo_total   = sum(b["saldo"] for b in blocos)
            # per-entidade (mesma entidade pode aparecer mais de uma vez → soma)
            porEnt = {}
            obs_global = []
            cat_global = {}
            for b in blocos:
                nome = b["nome_raw"]
                ent = porEnt.setdefault(nome, {
                    "pagamentoAutorizado": 0.0, "totalBancos": 0.0,
                    "observacoes": [], "categorias": {},
                })
                ent["pagamentoAutorizado"] += b["autorizado"]
                ent["totalBancos"]         += b["saldo"]
                ent["observacoes"].extend(b["observacoes"])
                for k, v in b["categorias"].items():
                    ent["categorias"][k] = ent["categorias"].get(k, 0.0) + v
                # global (consolidado da linha difPagAutorizado)
                for o in b["observacoes"]:
                    obs_global.append((nome, o["texto"], o["valor"]))
                for k, v in b["categorias"].items():
                    cat_global[k] = cat_global.get(k, 0.0) + v

            antigo = bancos.get(dd, {}).get("pagamentoAutorizado")
            ent_dd = bancos.setdefault(dd, {})
            ent_dd["pagamentoAutorizado"] = round(autoriz_total, 2)
            ent_dd["totalBancos"]         = round(saldo_total, 2)
            pagto = round(pagtos.get(dia, 0.0), 2)
            ent_dd["diferencasPagAutorizado"] = round(pagto + autoriz_total, 2)
            ent_dd["porEntidade"] = {
                k: {
                    "pagamentoAutorizado": round(v["pagamentoAutorizado"], 2),
                    "totalBancos":         round(v["totalBancos"], 2),
                    "observacoes":         v["observacoes"],
                    "categorias":          {kk: round(vv, 2) for kk, vv in v["categorias"].items()},
                } for k, v in porEnt.items()
            }

            # comentário automático na linha "Diferenças Pag Autorizado"
            cur_cmt = cmt_diff.get(dd)
            # Estrutura legacy do comentariosDFC: difPagAutorizado.DD = string (manual).
            # Marcação _auto vai num doc paralelo difPagAutorizado_auto.DD = true.
            auto_marks = (cmt_doc.get("difPagAutorizado_auto") or {})
            is_auto = bool(auto_marks.get(dd))
            if (not cur_cmt) or is_auto:
                linhas_cmt = []
                total_cmt = 0.0
                for nome in ("Tarifas", "Outros Pag.", "Pag. Tesouraria", "Transferência Protege"):
                    if nome in cat_global:
                        v = cat_global[nome]
                        linhas_cmt.append(f"{nome} — {_fmt(v)}")
                        total_cmt += v
                for ent_nome, txt, val in obs_global:
                    linhas_cmt.append(f"[{ent_nome}] {txt} — {_fmt(val)}")
                    total_cmt += val
                if linhas_cmt:
                    novo_texto = _texto_com_total(linhas_cmt, total_cmt)
                    cmt_updates[cmt_chave][f"difPagAutorizado.{dd}"] = novo_texto
                    cmt_updates[cmt_chave][f"difPagAutorizado_auto.{dd}"] = True

            relatorio.append((seg, ano, mes, dia, antigo, autoriz_total, saldo_total, pagto, len(porEnt)))

        doc["bancos"] = bancos
        if not dry_run:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(doc, f, ensure_ascii=False)

    if not dry_run:
        for docId, patches in cmt_updates.items():
            db.collection("comentariosDFC").document(docId).set(patches, merge=True)
            print(f"  ✓ comentariosDFC/{docId}: {len(patches)//2} dias com comentário auto")

    return relatorio


def main():
    ano_corr = dt.date.today().year
    mes_corr = dt.date.today().month
    dry      = "--dry-run" in sys.argv
    refresh  = "--refresh" in sys.argv
    ano = None; mes = None
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if len(args) >= 1: ano = int(args[0])
    if len(args) >= 2: mes = int(args[1])

    base = stage_local(refresh=refresh)
    if not base or not os.path.isdir(base):
        print(f"ERRO: conciliação não acessível ({CONC_MOUNT})")
        print("      (no WSL: mount -t drvfs '\\\\10.61.1.13\\cvl' /mnt/cvl)")
        sys.exit(1)

    print(f"Lendo conciliação de {base} {'(DRY-RUN)' if dry else ''}")
    por_dia = coleta(base, ano_filtro=ano, mes_filtro=mes)

    # auditoria de classificação
    audit = defaultdict(lambda: defaultdict(int))
    for (a, m, s), dias in por_dia.items():
        for d, blocos in dias.items():
            for b in blocos:
                audit[s][b["ent_norm"]] += 1
    print("\n=== Classificação de entidades (auditoria) ===")
    for seg in ("postos", "outras"):
        nomes = sorted(audit[seg])
        print(f"  {seg.upper()} ({len(nomes)}): {', '.join(nomes)}")

    rel = aplica(por_dia, dry_run=dry)

    print("\n=== Pagamento Autorizado + Total Bancos por dia ===")
    cur = None
    for seg, ano, mes, dia, antigo, autoriz, saldo, pagto, n_ent in rel:
        chave = (seg, ano, mes)
        if chave != cur:
            cur = chave
            print(f"\n[{seg.upper()} {ano}-{mes:02d}]  (pagtoDia | pagAutorizado | dif | totalBancos | nEnt)")
        a = f"{antigo:,.2f}" if isinstance(antigo, (int, float)) else "—"
        print(f"   dia {dia:02d}:  pagto {pagto:>14,.2f}  |  {a:>14} → {autoriz:>14,.2f}  |  dif {pagto+autoriz:>12,.2f}  |  TB {saldo:>14,.2f}  |  ent {n_ent}")

    print(f"\n{'(dry-run, nada gravado)' if dry else 'JSONs atualizados.'}  Linhas: {len(rel)}")


if __name__ == "__main__":
    main()
