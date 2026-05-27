"""ETL — DRE Postos. Lê 01 - DRE_POSTOS.xlsx (rede /mnt/controller/03 - POSTOS/
01 - Demonstrativos/Resultado/) e gera JSON consumido pela aba "Demonstrativo
de Resultado" quando segmento=postos.

Abas relevantes:
  - MÊS    → DRE mensal do mês atual × P01..P11 (11 lojas + total)
  - 2026   → DRE anual (12 meses × total consolidado)

Versão inicial: gera 1 arquivo consolidado por ano.

Output: /root/projeto_dre/dados_dre_postos/{ano}.json
Estrutura:
  {
    "ano": 2026, "geradoEm": "ISO", "v": 1,
    "linhas": [{nome, grupo, agrupamento, ordem, tipo}, ...],
    "lojas": ["P01", ..., "P11", "TOTAL"],
    "meses": [1..12],
    "dados": {
      "{linha}__{loja}__{mes}": valor,
      ...
    }
  }
"""
from __future__ import annotations
import os, json, datetime as dt
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

BASE = "/mnt/controller/03 - POSTOS/01 - Demonstrativos/Resultado"
F_DRE  = os.path.join(BASE, "01 - DRE_POSTOS.xlsx")
OUT_DIR = "/root/projeto_dre/dados_dre_postos"

ANO_ATUAL = 2026
POSTOS = [f"P{i:02d}" for i in range(1, 12)]   # P01..P11


def _to_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "."))
    except (ValueError, TypeError): return None


def _classificar_linha(nome: str, nome_normalizado: str):
    """Determina (grupo, agrupamento, tipo) a partir do nome da linha.
    grupo: header maior (Receita, Custo, Despesa, ...)
    agrupamento: bloco intermediário (Combustíveis, Despesas Administrativas, etc)
    tipo: section | subsection | item | total
    """
    n = nome.strip()
    nl = nome_normalizado.lower()
    # Headers de fórmula entre parênteses
    if n.startswith("( = )") or n.startswith("(=)"):
        return ("Total", n, "total")
    if n.startswith("(-)") or n.startswith("(- )"):
        return ("Total", n, "section")
    if n.startswith("Volume Vendido"):
        return ("Volume", n, "section")
    # Por palavras-chave no nome
    if "faturamento" in nl: return ("Receita", n, "section")
    if "custo" in nl and "total" in nl: return ("Custo", n, "section")
    if "lucro" in nl: return ("Resultado", n, "total")
    if "despesa" in nl: return ("Despesa", n, "section")
    if "varia" in nl and "estoque" in nl: return ("Resultado", n, "item")
    # default = item
    return ("Outros", n, "item")


def parse_aba_mes(wb):
    """Extrai DRE da aba MÊS (P01..P11 + TOTAL). Cabeçalho na linha 18,
    dados a partir da linha 19. Cada loja ocupa 2 colunas (Valor, Part%).
    Retorna lista de tuples (linha_id, linha_nome, loja, valor, part_pct).
    """
    ws = wb["MÊS"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 19: return []
    header = rows[18]
    # Mapa col→loja (col com header == "P0X" ou "TOTAL")
    col_loja = {}
    for ci, c in enumerate(header):
        s = str(c).strip() if c else ""
        if s in POSTOS or s == "TOTAL":
            col_loja[ci] = s
    out = []
    for ri in range(19, len(rows)):
        r = rows[ri]
        nome = r[1] if len(r) > 1 else None
        if not nome or not str(nome).strip(): continue
        nome = str(nome).strip()
        if nome.startswith("#"): continue
        for ci, loja in col_loja.items():
            if ci >= len(r): continue
            valor = _to_float(r[ci])
            part = _to_float(r[ci + 1]) if ci + 1 < len(r) else None
            if valor is None and part is None: continue
            out.append((ri, nome, loja, valor, part))
    return out


def parse_aba_ano(wb, ano):
    """Extrai DRE da aba do ano (ex: 2026): 12 meses como colunas, total
    consolidado. Cabeçalho na linha 3 (1-2026, 2-2026, ..., 12-2026).
    Retorna lista de tuples (linha_id, linha_nome, mes, valor, part_pct).
    """
    sheet_name = str(ano)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4: return []
    header = rows[3]
    col_mes = {}
    for ci, c in enumerate(header):
        s = str(c).strip() if c else ""
        # Format: "1-2026", "2-2026", etc.
        if s and "-" in s:
            try:
                mes_str, ano_str = s.split("-", 1)
                if int(ano_str) == ano:
                    col_mes[ci] = int(mes_str)
            except (ValueError, TypeError):
                pass
    out = []
    for ri in range(4, len(rows)):
        r = rows[ri]
        nome = r[1] if len(r) > 1 else None
        if not nome or not str(nome).strip(): continue
        nome = str(nome).strip()
        if nome.startswith("#"): continue
        for ci, mes in col_mes.items():
            if ci >= len(r): continue
            valor = _to_float(r[ci])
            part = _to_float(r[ci + 1]) if ci + 1 < len(r) else None
            if valor is None and part is None: continue
            out.append((ri, nome, mes, valor, part))
    return out


def main():
    if not os.path.exists(F_DRE):
        raise FileNotFoundError(f"{F_DRE} não encontrado. Confira o mount de /mnt/controller.")

    os.makedirs(OUT_DIR, exist_ok=True)
    print(f"→ Lendo {F_DRE}…")
    wb = load_workbook(F_DRE, read_only=True, data_only=True)

    # 1) Aba MÊS (lojas × DRE) — mês atual
    print("→ Parse aba MÊS (lojas × DRE)…")
    rows_mes = parse_aba_mes(wb)
    print(f"   {len(rows_mes)} células lidas (linha × loja)")

    # 2) Aba ANO (12 meses × DRE consolidado)
    print(f"→ Parse aba {ANO_ATUAL} (12 meses × consolidado)…")
    rows_ano = parse_aba_ano(wb, ANO_ATUAL)
    print(f"   {len(rows_ano)} células lidas (linha × mês)")

    # Coleta linhas únicas (preserva ordem da aba ANO se disponível, senão MÊS)
    linhas_seen = {}
    linhas_ordem = []
    for tup in rows_ano + rows_mes:
        ri, nome = tup[0], tup[1]
        if nome not in linhas_seen:
            grupo, agrup, tipo = _classificar_linha(nome, nome)
            linhas_seen[nome] = {"nome": nome, "grupo": grupo, "agrupamento": agrup,
                                  "ordem": ri, "tipo": tipo}
            linhas_ordem.append(nome)

    # Monta dados — chave "linha__loja__mes"
    dados = {}
    # Da aba ANO: mês conhecido, loja = TOTAL
    for ri, nome, mes, valor, part in rows_ano:
        if valor is None: continue
        dados[f"{nome}__TOTAL__{mes}"] = round(valor, 2)
        if part is not None:
            dados[f"{nome}__TOTAL__{mes}__pct"] = round(part, 6)
    # Da aba MÊS: loja conhecida; mês inferido pela data corrente do arquivo
    #   (planilha é "mês atual" — vou marcar com mes=__MES_ATUAL__ pra UI tratar)
    # Tentativa: detecta o mês via aba "MÊS" (a planilha tem alguma célula com data)
    mes_atual = None
    try:
        ws_mes = wb["MÊS"]
        # Pega data da linha 1 col C (vista anteriormente)
        rows_mes_raw = list(ws_mes.iter_rows(values_only=True))
        for cand_row in rows_mes_raw[:3]:
            for c in cand_row:
                if isinstance(c, dt.datetime):
                    mes_atual = c.month
                    break
            if mes_atual: break
    except Exception:
        pass
    if mes_atual:
        print(f"   Mês detectado na aba MÊS: {mes_atual}")
        for ri, nome, loja, valor, part in rows_mes:
            if valor is None: continue
            dados[f"{nome}__{loja}__{mes_atual}"] = round(valor, 2)
            if part is not None:
                dados[f"{nome}__{loja}__{mes_atual}__pct"] = round(part, 6)
    wb.close()

    payload = {
        "ano": ANO_ATUAL,
        "geradoEm": dt.datetime.now().isoformat(timespec="seconds"),
        "v": 1,
        "mesAtual": mes_atual,
        "linhas": list(linhas_seen.values()),
        "lojas": POSTOS + ["TOTAL"],
        "meses": list(range(1, 13)),
        "dados": dados,
    }
    out_path = os.path.join(OUT_DIR, f"{ANO_ATUAL}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    print(f"\nOK gerado: {out_path}  ({os.path.getsize(out_path):,} bytes)")
    print(f"  Linhas DRE: {len(linhas_seen)}")
    print(f"  Células de dados: {len(dados)}")


if __name__ == "__main__":
    main()
